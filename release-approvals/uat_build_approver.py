"""
UAT Build Approver
------------------
Reads the latest uat_release_report_*.json, polls each build until it reaches
the "UAT Approval" gate, then defers it to a user-specified date/time via the
Azure DevOps REST API. Exports an Excel + JSON approval report.

SETUP (one time):
  pip install requests openpyxl

CONFIGURE: Fill in the CONFIG section below, then run:
  python uat_build_approver.py

You will be prompted for the defer date/time and timezone at startup.
"""

import concurrent.futures
import glob
import json
import os
import sys
import time
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import common

# Reports are written to <project-root>/reports/uat/ regardless of where the script is run from.
# Input JSON (from uat_build_tracker.py) is also found there automatically.
_REPORTS_DIR = os.path.normpath(
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "reports", "uat")
)

# ─────────────────────────────────────────────
#  CONFIG  ← Fill these in before running
# ─────────────────────────────────────────────
CONFIG = {
    "org":                   "scmdevops",
    "project":               "Leo.TPRM",
    "pat":                   "<YOUR_PAT_HERE>",
    "approval_stage_name":   "UAT Approval",
    "defer_comment":         "Deferred by uat_build_approver.py",
    # Leave empty to auto-detect the latest uat_release_report_*.json in output_dir
    "input_json":            "",
    "poll_interval_seconds": 60,
    "poll_timeout_minutes":  60,
    "output_dir":            _REPORTS_DIR,
}
# ─────────────────────────────────────────────

common.load_credentials(__file__, CONFIG)
_session, BASE_URL = common.build_session(CONFIG)
get, post, patch = common.make_http_fns(_session)

# ── State constants ────────────────────────────────────────────────────────────
STATE_ORDER = {
    "deferred": 0,
    "build_failed": 1,
    "timed_out": 2,
    "no_approval_found": 3,
    "api_error": 4,
    "pending": 5,
}
STATE_LABELS = {
    "deferred":           "Deferred",
    "build_failed":       "Build Failed",
    "timed_out":          "Timed Out",
    "no_approval_found":  "No Gate Found",
    "api_error":          "API Error",
    "pending":            "Still Pending",
}
STATE_EMOJI = {
    "deferred":           "✅",
    "build_failed":       "❌",
    "timed_out":          "⏱",
    "no_approval_found":  "⚠️",
    "api_error":          "❌",
    "pending":            "...",
}


# ── Defer time prompt ──────────────────────────────────────────────────────────

def parse_defer_input(time_str, tz_name):
    """
    Parse a user-supplied time string into a UTC datetime.
    Accepts: 'tomorrow HH:MM' or 'YYYY-MM-DD HH:MM'
    """
    tz = ZoneInfo(tz_name)
    now_local = datetime.now(tz)
    time_str = time_str.strip()

    if time_str.lower().startswith("tomorrow"):
        hhmm = time_str[len("tomorrow"):].strip() or "00:00"
        t = datetime.strptime(hhmm, "%H:%M")
        local_dt = (now_local + timedelta(days=1)).replace(
            hour=t.hour, minute=t.minute, second=0, microsecond=0
        )
    else:
        local_dt = datetime.strptime(time_str, "%Y-%m-%d %H:%M").replace(tzinfo=tz)

    return local_dt.astimezone(ZoneInfo("UTC"))


def prompt_defer_time():
    """Interactively prompt the user for defer date/time and timezone. Returns UTC datetime."""
    print("\n  ── Defer Time Setup ─────────────────────────────────────────")
    print("  Examples: 'tomorrow 07:00'  |  '2026-03-08 07:00'")
    print()

    while True:
        try:
            time_str = input("  Enter date/time  : ").strip()
            tz_str   = input("  Enter timezone   (IANA name) [Asia/Kolkata]: ").strip() or "Asia/Kolkata"

            try:
                defer_utc = parse_defer_input(time_str, tz_str)
            except ZoneInfoNotFoundError:
                print(f"\n  Unknown timezone: '{tz_str}'. Use an IANA name like 'Asia/Kolkata', 'UTC', 'America/New_York'.\n")
                continue
            except ValueError as e:
                print(f"\n  Could not parse '{time_str}': {e}. Try 'tomorrow 07:00' or '2026-03-08 07:00'.\n")
                continue

            tz        = ZoneInfo(tz_str)
            local_dt  = defer_utc.astimezone(tz)
            tz_abbr   = local_dt.strftime("%Z")

            print(f"\n  Will defer to: {local_dt.strftime('%Y-%m-%d %H:%M:%S')} {tz_abbr}"
                  f"  =  {defer_utc.strftime('%Y-%m-%d %H:%M:%S')} UTC")

            confirm = input("  Confirm? [Y/n]: ").strip().lower()
            if confirm in ("", "y", "yes"):
                print()
                return defer_utc, local_dt.strftime(f"%Y-%m-%d %H:%M {tz_abbr}")
            print()

        except (EOFError, KeyboardInterrupt):
            print("\n\n  Cancelled by user.")
            os._exit(0)


# ── Input loading ──────────────────────────────────────────────────────────────

def resolve_input_json():
    if CONFIG["input_json"]:
        return CONFIG["input_json"]
    pattern = os.path.join(CONFIG["output_dir"], "uat_release_report_*.json")
    matches = sorted(glob.glob(pattern))
    if not matches:
        raise FileNotFoundError(
            f"No uat_release_report_*.json found in '{CONFIG['output_dir']}'.\n"
            "Run uat_build_tracker.py first."
        )
    path = matches[-1]
    print(f"  Using input file: {path}")
    return path


def load_builds(json_path):
    with open(json_path, encoding="utf-8") as f:
        entries = json.load(f)
    valid = [e for e in entries if e.get("status") == "Found" and e.get("build_id")]
    skipped = len(entries) - len(valid)
    if skipped:
        print(f"  Skipped {skipped} entries without a build (status != Found).")
    print(f"  Loaded {len(valid)} build(s) to defer.")
    return valid


# ── Azure DevOps polling & deferral ───────────────────────────────────────────

def get_build_status(build_id):
    try:
        return get(f"{BASE_URL}/build/builds/{build_id}")
    except Exception as e:
        return {"status": "error", "result": None, "_error": str(e)}


def get_build_timeline(build_id):
    try:
        return get(f"{BASE_URL}/build/builds/{build_id}/timeline").get("records", [])
    except Exception:
        return []


def find_approval_checkpoint(records, stage_name):
    """
    Find the first timeline Checkpoint record that is waiting/pending/deferred
    and whose name contains stage_name (case-insensitive substring).
    Including 'deferred' allows re-deferral by re-running the script with a new time.
    """
    stage_lower = stage_name.lower()
    for r in records:
        if (
            r.get("type") == "Checkpoint"
            and r.get("state", "").lower() in ("waiting", "pending", "deferred")
            and stage_lower in r.get("name", "").lower()
        ):
            return r
    return None


def defer_checkpoint(approval_id, pipeline_name, defer_utc):
    """Defer the given checkpoint to defer_utc. Returns True on success."""
    defer_str = defer_utc.strftime("%Y-%m-%dT%H:%M:%S.000Z")
    try:
        patch(f"{BASE_URL}/pipelines/approvals", [
            {
                "approvalId": approval_id,
                "status":     "deferred",
                "deferredTo": defer_str,
                "comment":    CONFIG["defer_comment"],
            }
        ])
        print(f"    [{pipeline_name}] Deferred checkpoint {approval_id} → {defer_str}")
        return True
    except requests.HTTPError as e:
        code = e.response.status_code if e.response is not None else "?"
        try:
            detail = e.response.json().get("message", str(e)) if e.response is not None else str(e)
        except Exception:
            detail = str(e)
        print(f"    [{pipeline_name}] FAILED to defer — HTTP {code}: {detail[:120]}")
        return False
    except Exception as e:
        print(f"    [{pipeline_name}] FAILED to defer — {e}")
        return False


# ── Per-build poll logic ───────────────────────────────────────────────────────

def poll_and_defer_build(entry, states, start_times, defer_utc):
    build_id      = entry["build_id"]
    pipeline_name = entry["pipeline_name"]
    state         = states[build_id]

    if state["state"] != "pending":
        return

    elapsed = (datetime.now(timezone.utc) - start_times[build_id]).total_seconds()
    if elapsed > CONFIG["poll_timeout_minutes"] * 60:
        state["state"] = "timed_out"
        return

    build_resp = get_build_status(build_id)

    if "_error" in build_resp:
        state["state"] = "api_error"
        state["error"] = build_resp["_error"]
        return

    status = build_resp.get("status", "")
    result = build_resp.get("result")

    if status == "completed":
        state["state"] = "no_approval_found" if result == "succeeded" else "build_failed"
        if result != "succeeded":
            state["result"] = result or "unknown"
        return

    records    = get_build_timeline(build_id)
    checkpoint = find_approval_checkpoint(records, CONFIG["approval_stage_name"])

    if checkpoint:
        success = defer_checkpoint(checkpoint["id"], pipeline_name, defer_utc)
        if success:
            state["state"]       = "deferred"
            state["actioned_at"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
        else:
            state["state"] = "api_error"
            state["error"] = f"Defer PATCH failed for checkpoint {checkpoint['id']}"


# ── Poll loop ──────────────────────────────────────────────────────────────────

def print_status_table(entries, states, cycle):
    print(f"\n  {'─'*66}")
    print(f"  Cycle {cycle} — {datetime.now(timezone.utc).strftime('%H:%M:%S UTC')}")
    print(f"  {'─'*66}")
    for entry in entries:
        sid   = entry["build_id"]
        name  = entry["pipeline_name"][:38].ljust(38)
        s     = states[sid]["state"]
        emoji = STATE_EMOJI.get(s, "?")
        label = STATE_LABELS.get(s, s)
        print(f"  {name}  {emoji}  {label}")
    print(f"  {'─'*66}")
    pending  = sum(1 for e in entries if states[e["build_id"]]["state"] == "pending")
    deferred = sum(1 for e in entries if states[e["build_id"]]["state"] == "deferred")
    failed   = sum(1 for e in entries if states[e["build_id"]]["state"] == "build_failed")
    timedout = sum(1 for e in entries if states[e["build_id"]]["state"] == "timed_out")
    print(f"  Pending: {pending}  |  Deferred: {deferred}  |  Failed: {failed}  |  Timed Out: {timedout}\n")


def run_poll_loop(entries, defer_utc):
    states = {
        e["build_id"]: {"state": "pending", "actioned_at": None, "error": None, "result": None}
        for e in entries
    }
    start_times = {e["build_id"]: datetime.now(timezone.utc) for e in entries}

    print(
        f"  Polling {len(entries)} build(s) — "
        f"every {CONFIG['poll_interval_seconds']}s, "
        f"timeout {CONFIG['poll_timeout_minutes']}min\n"
    )

    cycle = 0
    try:
        while any(s["state"] == "pending" for s in states.values()):
            cycle += 1
            pending_entries = [e for e in entries if states[e["build_id"]]["state"] == "pending"]
            workers = min(len(pending_entries), 8)
            with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as ex:
                futures = {
                    ex.submit(poll_and_defer_build, entry, states, start_times, defer_utc): entry
                    for entry in pending_entries
                }
                concurrent.futures.wait(futures)

            print_status_table(entries, states, cycle)

            if any(s["state"] == "pending" for s in states.values()):
                time.sleep(CONFIG["poll_interval_seconds"])
    except KeyboardInterrupt:
        print("\n\n  Interrupted by user — exiting.")
        os._exit(1)

    return states


# ── Result merging ─────────────────────────────────────────────────────────────

def build_final_results(entries, states, defer_label):
    results = []
    for entry in entries:
        bid   = entry["build_id"]
        state = states[bid]
        results.append({
            "pipeline_name": entry.get("pipeline_name", ""),
            "status":        state["state"],
            "build_number":  entry.get("build_number", "—"),
            "branch":        entry.get("branch", "—"),
            "actioned_at":   state.get("actioned_at") or "—",
            "deferred_to":   defer_label if state["state"] == "deferred" else "—",
            "notes":         state.get("error") or state.get("result") or "",
            "build_url":     entry.get("build_url", ""),
            "build_id":      bid,
        })
    results.sort(key=lambda r: (STATE_ORDER.get(r["status"], 99), r["pipeline_name"]))
    return results


# ── Excel report ───────────────────────────────────────────────────────────────

def build_excel_report(results, defer_label):
    wb = Workbook()

    from common import (BLUE_DARK, BLUE_MID, BLUE_LIGHT, GREEN_BG, AMBER_BG, RED_BG,
                        WHITE, GREY_ROW, header_font, cell_font, fill, thin_border, center, left)

    ws = wb.active
    ws.title = "UAT Approval Report"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    ws["A1"] = f"UAT Release — UAT Build Approver  |  Deferred to: {defer_label}"
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color=WHITE)
    ws["A1"].fill = fill(BLUE_DARK)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    ws["A2"] = (
        f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
        f"  |  Org: {CONFIG['org']}  |  Project: {CONFIG['project']}"
    )
    ws["A2"].font = Font(name="Arial", size=9, color=WHITE)
    ws["A2"].fill = fill(BLUE_MID)
    ws["A2"].alignment = center()
    ws.row_dimensions[2].height = 18

    total    = len(results)
    deferred = sum(1 for r in results if r["status"] == "deferred")
    failed   = sum(1 for r in results if r["status"] == "build_failed")
    timedout = sum(1 for r in results if r["status"] == "timed_out")

    ws.merge_cells("A3:B3"); ws["A3"] = "Total Builds"
    ws.merge_cells("C3:D3"); ws["C3"] = "Deferred"
    ws.merge_cells("E3:F3"); ws["E3"] = "Build Failed"
    ws.merge_cells("G3:H3"); ws["G3"] = "Timed Out"
    ws.merge_cells("A4:B4"); ws["A4"] = total
    ws.merge_cells("C4:D4"); ws["C4"] = deferred
    ws.merge_cells("E4:F4"); ws["E4"] = failed
    ws.merge_cells("G4:H4"); ws["G4"] = timedout

    for col_letter, bg in [("A", BLUE_LIGHT), ("C", GREEN_BG), ("E", RED_BG), ("G", AMBER_BG)]:
        for row in [3, 4]:
            cell = ws[f"{col_letter}{row}"]
            cell.fill = fill(bg)
            cell.alignment = center()
            cell.font = cell_font(9, bold=True) if row == 3 else Font(name="Arial", size=20, bold=True)

    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 40
    ws.row_dimensions[5].height = 8

    headers    = ["Repo / Pipeline", "Status", "Build #", "Branch", "Actioned At", "Deferred To", "Notes", "Build Link"]
    col_widths = [35, 20, 14, 20, 22, 26, 30, 14]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=6, column=col_idx, value=h)
        cell.font = header_font()
        cell.fill = fill(BLUE_MID)
        cell.alignment = center()
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[6].height = 22

    STATUS_FILL = {
        "deferred":          (fill(GREEN_BG), "276221"),
        "build_failed":      (fill(RED_BG),   "8B0000"),
        "timed_out":         (fill(AMBER_BG), "7D5200"),
        "no_approval_found": (fill(AMBER_BG), "7D5200"),
        "api_error":         (fill(RED_BG),   "8B0000"),
        "pending":           (fill(AMBER_BG), "7D5200"),
    }

    for row_idx, r in enumerate(results, start=7):
        bg = GREY_ROW if row_idx % 2 == 0 else WHITE
        status = r["status"]
        s_fill, s_color = STATUS_FILL.get(status, (fill(RED_BG), "8B0000"))
        link_label = "Open ↗" if status in ("deferred", "no_approval_found") else "—"

        row_data = [
            r["pipeline_name"],
            STATE_LABELS.get(status, status),
            r["build_number"],
            r["branch"],
            r["actioned_at"],
            r["deferred_to"],
            r["notes"],
            link_label,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border()
            cell.alignment = left() if col_idx in (1, 7) else center()
            cell.font = cell_font()

            if col_idx == 2:
                cell.fill = s_fill
                cell.font = Font(name="Arial", size=10, bold=True, color=s_color)
            else:
                cell.fill = fill(bg)

            if col_idx == 8 and r.get("build_url") and status in ("deferred", "no_approval_found"):
                cell.hyperlink = r["build_url"]
                cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")

        ws.row_dimensions[row_idx].height = 18

    ws.freeze_panes = "A7"

    # ── Sheet 2: How To Use ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("How To Use")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 90

    instructions = [
        ("UAT Build Approver — How To Use", BLUE_DARK, 14, True),
        ("", WHITE, 10, False),
        ("PURPOSE", BLUE_MID, 11, True),
        ("  Run this script before a UAT release to defer the 'UAT Approval' gate for each build", WHITE, 10, False),
        ("  to a specific date and time. The approver can review the Excel report for confirmation.", WHITE, 10, False),
        ("", WHITE, 10, False),
        ("WORKFLOW", BLUE_MID, 11, True),
        ("  1. Run uat_build_tracker.py → finds QC-passed builds, saves uat_release_report_*.json", WHITE, 10, False),
        ("  2. Run uat_build_approver.py → polls builds, defers UAT Approval gate, saves report", WHITE, 10, False),
        ("  You will be prompted for the defer date/time and timezone when the script starts.", WHITE, 10, False),
        ("", WHITE, 10, False),
        ("RE-DEFERRING TO A NEW TIME", BLUE_MID, 11, True),
        ("  Re-run the script and enter the new defer time. It will update already-deferred", WHITE, 10, False),
        ("  approvals automatically (no manual steps needed).", WHITE, 10, False),
        ("", WHITE, 10, False),
        ("STEP 1 — Configure the script", BLUE_MID, 11, True),
        ("  • pat                  → Full-access PAT (Build: Read + Pipelines: Read & Manage)", WHITE, 10, False),
        ("  • approval_stage_name  → Name of the gate (default: 'UAT Approval')", WHITE, 10, False),
        ("  • input_json           → Leave empty to auto-detect the latest report file", WHITE, 10, False),
        ("  • poll_interval_seconds→ How often to re-check builds (default: 60)", WHITE, 10, False),
        ("  • poll_timeout_minutes → Give up after this many minutes (default: 60)", WHITE, 10, False),
        ("", WHITE, 10, False),
        ("STEP 2 — Run the script", BLUE_MID, 11, True),
        ("  python uat_build_approver.py", WHITE, 10, False),
        ("  When prompted: enter time like 'tomorrow 07:00' or '2026-03-08 07:00'", WHITE, 10, False),
        ("  Timezone: use an IANA name like 'Asia/Kolkata', 'UTC', 'America/New_York'", WHITE, 10, False),
        ("", WHITE, 10, False),
        ("UNDERSTANDING THE STATUS COLOURS", BLUE_MID, 11, True),
        ("  Deferred        → Gate found and deferred to the specified time", WHITE, 10, False),
        ("  Build Failed    → Build completed with a failure before reaching the gate", WHITE, 10, False),
        ("  Timed Out       → Gate was not reached within poll_timeout_minutes", WHITE, 10, False),
        ("  No Gate Found   → Build completed successfully but no approval gate was seen", WHITE, 10, False),
        ("  API Error       → HTTP or network error during polling or deferral", WHITE, 10, False),
    ]

    for i, (text, bg_hex, font_size, bold) in enumerate(instructions, start=1):
        cell = ws2.cell(row=i, column=1, value=text)
        cell.font = Font(name="Arial", size=font_size, bold=bold,
                         color=WHITE if bg_hex != WHITE else "000000")
        cell.fill = fill(bg_hex) if bg_hex != WHITE else PatternFill()
        cell.alignment = left()
        ws2.row_dimensions[i].height = 18

    return wb


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print(f"\n{'='*55}")
    print("  UAT Build Approver — Azure DevOps")
    print(f"{'='*55}")
    print(f"  Org:     {CONFIG['org']}")
    print(f"  Project: {CONFIG['project']}")
    print(f"  Gate:    '{CONFIG['approval_stage_name']}' (will be DEFERRED)")
    print(f"  Timeout: {CONFIG['poll_timeout_minutes']} min  |  Interval: {CONFIG['poll_interval_seconds']} s")
    print(f"{'='*55}")

    defer_utc, defer_label = prompt_defer_time()

    try:
        json_path = resolve_input_json()
    except FileNotFoundError as e:
        print(f"  ERROR: {e}")
        return

    entries = load_builds(json_path)
    if not entries:
        print("  No builds found — nothing to defer.")
        return

    print(f"\n  Starting poll loop for {len(entries)} build(s)...\n")
    states  = run_poll_loop(entries, defer_utc)
    results = build_final_results(entries, states, defer_label)

    output_dir = CONFIG.get("output_dir", ".")
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
    xlsx_file = os.path.join(output_dir, f"uat_approval_report_{timestamp}.xlsx")
    json_file = os.path.join(output_dir, f"uat_approval_report_{timestamp}.json")

    print("  Building Excel report...")
    wb = build_excel_report(results, defer_label)
    wb.save(xlsx_file)

    print("  Saving JSON...")
    with open(json_file, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, default=str)

    deferred          = sum(1 for r in results if r["status"] == "deferred")
    build_failed      = sum(1 for r in results if r["status"] == "build_failed")
    timed_out         = sum(1 for r in results if r["status"] == "timed_out")
    no_approval_found = sum(1 for r in results if r["status"] == "no_approval_found")
    api_errors        = sum(1 for r in results if r["status"] == "api_error")

    print(f"\n{'='*55}")
    print(f"  Reports saved:")
    print(f"       Excel: {xlsx_file}")
    print(f"       JSON:  {json_file}")
    print(f"{'='*55}")
    print(f"  Builds processed  : {len(results)}")
    print(f"  Deferred          : {deferred}  →  {defer_label}")
    print(f"  Build Failed      : {build_failed}")
    print(f"  Timed Out         : {timed_out}")
    print(f"  No Gate Found     : {no_approval_found}")
    print(f"  API Errors        : {api_errors}")
    print(f"{'='*55}\n")

    if build_failed or timed_out or api_errors:
        print("  Builds that were not deferred:")
        for r in results:
            if r["status"] not in ("deferred", "no_approval_found"):
                print(f"    - {r['pipeline_name']}: {STATE_LABELS.get(r['status'], r['status'])} — {r['notes']}")
        print()


if __name__ == "__main__":
    main()
