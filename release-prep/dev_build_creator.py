"""
DEV Build Creator
-----------------
For each pipeline in the project, triggers a new build on the DEVELOPMENT branch.
Exports a ready-to-share Excel report so the approver can review and approve each build.

SETUP (one time):
  pip install requests openpyxl

NOTE: Your PAT must have Build: Read AND Build: Execute (Queue) permissions.

CONFIGURE: Fill in the CONFIG section below, then run:
  python dev_build_creator.py
"""

import concurrent.futures
import os
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import json
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Reports are written to <project-root>/reports/qc/ regardless of where the script is run from.
_REPORTS_DIR = os.path.normpath(
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "reports", "qc")
)

# ─────────────────────────────────────────────
#  CONFIG  ← Fill these in before running
# ─────────────────────────────────────────────
CONFIG = {
    "org":        "scmdevops",         # Azure DevOps org name
    "project":    "Leo.TPRM",          # Project name
    "pat":        "<YOUR_PAT_HERE>",  # PAT with Build: Read + Execute
    # Branch to trigger builds on
    "dev_branch": "DEVELOPMENT",
    # Leave pipeline_names as [] to auto-discover all pipelines in the project.
    # Or list specific names to trigger only those pipelines.
    "pipeline_names": [],
    # Pipelines whose name contains any of these substrings (case-insensitive) are skipped.
    "exclude_pipeline_name_patterns": ["-CI", "plugin", "automation", "NotInUse", "Not_in_use", "Cypress"],
    # Directory where Excel and JSON reports are saved. Created automatically if it doesn't exist.
    "output_dir": _REPORTS_DIR,
}
# ─────────────────────────────────────────────

# ── Load shared credentials from config.json at the project root ────────────
_CONFIG_FILE = os.path.normpath(
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "config.json")
)
try:
    with open(_CONFIG_FILE, encoding="utf-8") as _f:
        _shared = json.load(_f)
    CONFIG["org"]     = _shared.get("org",     CONFIG["org"])
    CONFIG["project"] = _shared.get("project", CONFIG["project"])
    CONFIG["pat"]     = _shared.get("pat",      CONFIG["pat"])
except FileNotFoundError:
    pass  # Falls back to the values set in CONFIG above
# ────────────────────────────────────────────────────────────────────────────

BASE_URL = f"https://dev.azure.com/{CONFIG['org']}/{CONFIG['project']}/_apis"
AUTH = ("", CONFIG["pat"])
HEADERS = {"Content-Type": "application/json"}

_session = requests.Session()
_session.auth = AUTH
_session.headers.update(HEADERS)
_retry = Retry(total=3, backoff_factor=0.5, status_forcelist=[429, 500, 502, 503, 504])
_adapter = HTTPAdapter(pool_connections=20, pool_maxsize=20, max_retries=_retry)
_session.mount("https://", _adapter)


def get(url, params=None):
    params = params or {}
    params["api-version"] = "7.1"
    r = _session.get(url, params=params, timeout=(10, 30))
    r.raise_for_status()
    return r.json()


def post(url, body, params=None):
    params = params or {}
    params["api-version"] = "7.1"
    r = _session.post(url, json=body, params=params, timeout=(10, 30))
    r.raise_for_status()
    return r.json()


def get_all_pipelines():
    """Fetch all pipelines in the project, applying name filters from CONFIG."""
    data = get(f"{BASE_URL}/pipelines", {"$top": 500})
    pipelines = data.get("value", [])

    if CONFIG["pipeline_names"]:
        pipelines = [p for p in pipelines if p["name"] in CONFIG["pipeline_names"]]
    else:
        exclude_patterns = [p.lower() for p in CONFIG.get("exclude_pipeline_name_patterns", [])]
        if exclude_patterns:
            pipelines = [
                p for p in pipelines
                if not any(pat in p["name"].lower() for pat in exclude_patterns)
            ]

    return pipelines


def trigger_dev_build(pipeline_id, pipeline_name):
    """
    Queue a new build for the given pipeline on the configured dev branch.
    Returns a result dict with status, build number, and a direct link.
    """
    branch_ref = f"refs/heads/{CONFIG['dev_branch']}"
    try:
        print(f"    [{pipeline_name}] triggering build on '{CONFIG['dev_branch']}'...")
        response = post(f"{BASE_URL}/build/builds", {
            "definition": {"id": pipeline_id},
            "sourceBranch": branch_ref,
        })

        build_id     = response.get("id")
        build_number = response.get("buildNumber", "N/A")
        queued_on    = response.get("queueTime", "")
        queued_dt    = datetime.strptime(queued_on[:19], "%Y-%m-%dT%H:%M:%S") if queued_on else None
        requested_by = response.get("requestedFor", {}).get("displayName", "N/A")
        build_url    = (
            f"https://{CONFIG['org']}.visualstudio.com/{CONFIG['project']}"
            f"/_build/results?buildId={build_id}"
        )

        return {
            "status":       "Triggered",
            "build_id":     build_id,
            "build_number": build_number,
            "branch":       CONFIG["dev_branch"],
            "requested_by": requested_by,
            "queued_at":    queued_dt.strftime("%Y-%m-%d %H:%M UTC") if queued_dt else "N/A",
            "build_url":    build_url,
        }

    except requests.HTTPError as e:
        status_code = e.response.status_code if e.response is not None else "?"
        try:
            detail = e.response.json().get("message", str(e)) if e.response is not None else str(e)
        except Exception:
            detail = str(e)

        # 400: branch doesn't exist in this pipeline / pipeline disabled
        # 403: PAT lacks Queue permission
        if status_code == 400:
            reason = f"Branch not found or pipeline disabled ({detail[:80]})"
        elif status_code == 403:
            reason = "Permission denied — PAT needs Build: Execute (Queue)"
        else:
            reason = f"HTTP {status_code}: {detail[:80]}"

        print(f"    [{pipeline_name}] FAILED — {reason}")
        return {"status": "Failed to Trigger", "error": reason}

    except Exception as e:
        print(f"    [{pipeline_name}] ERROR — {e}")
        return {"status": "Failed to Trigger", "error": str(e)}


def build_excel_report(results):
    wb = Workbook()

    # ── Styles ──────────────────────────────────────────────────────────────
    BLUE_DARK  = "1F3864"
    BLUE_MID   = "2E75B6"
    BLUE_LIGHT = "D6E4F0"
    GREEN_BG   = "E2EFDA"
    AMBER_BG   = "FFF2CC"
    RED_BG     = "FFDDD8"
    WHITE      = "FFFFFF"
    GREY_ROW   = "F5F7FA"

    def header_font(size=11, bold=True, color=WHITE):
        return Font(name="Arial", size=size, bold=bold, color=color)

    def cell_font(size=10, bold=False, color="000000"):
        return Font(name="Arial", size=size, bold=bold, color=color)

    def fill(hex_color):
        return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

    def thin_border():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def left():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ── Sheet 1: Summary Dashboard ────────────────────────────────────────
    ws = wb.active
    ws.title = "DEV Build Report"
    ws.sheet_view.showGridLines = False

    # Title block
    ws.merge_cells("A1:H1")
    ws["A1"] = f"QC Release — DEV Build Creator  |  Branch: {CONFIG['dev_branch']}"
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color=WHITE)
    ws["A1"].fill = fill(BLUE_DARK)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    ws["A2"] = (
        f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
        f"  |  Org: {CONFIG['org']}  |  Project: {CONFIG['project']}"
    )
    ws["A2"].font = Font(name="Arial", size=9, color=WHITE)
    ws["A2"].fill = fill(BLUE_MID)
    ws["A2"].alignment = center()
    ws.row_dimensions[2].height = 18

    # KPI row
    total     = len(results)
    triggered = sum(1 for r in results if r["status"] == "Triggered")
    failed    = sum(1 for r in results if r["status"] == "Failed to Trigger")

    ws.merge_cells("A3:B3"); ws["A3"] = "Total Pipelines"
    ws.merge_cells("C3:E3"); ws["C3"] = "Builds Triggered"
    ws.merge_cells("F3:H3"); ws["F3"] = "Failed to Trigger"
    ws.merge_cells("A4:B4"); ws["A4"] = total
    ws.merge_cells("C4:E4"); ws["C4"] = triggered
    ws.merge_cells("F4:H4"); ws["F4"] = failed

    for col_letter, bg in [("A", BLUE_LIGHT), ("C", GREEN_BG), ("F", RED_BG)]:
        for row in [3, 4]:
            cell = ws[f"{col_letter}{row}"]
            cell.fill = fill(bg)
            cell.alignment = center()
            if row == 3:
                cell.font = cell_font(9, bold=True)
            else:
                cell.font = Font(name="Arial", size=20, bold=True)

    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 40
    ws.row_dimensions[5].height = 8  # spacer

    # Column headers
    headers    = ["Repo / Pipeline", "Status", "Build #", "Branch", "Triggered By", "Queued At", "Notes", "Build Link"]
    col_widths = [35, 20, 14, 22, 24, 22, 30, 14]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=6, column=col_idx, value=h)
        cell.font = header_font()
        cell.fill = fill(BLUE_MID)
        cell.alignment = center()
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[6].height = 22

    # Data rows
    for row_idx, r in enumerate(results, start=7):
        bg     = GREY_ROW if row_idx % 2 == 0 else WHITE
        status = r["status"]

        if status == "Triggered":
            status_fill  = fill(GREEN_BG)
            status_color = "276221"
            link_label   = "Open ↗"
            notes        = ""
        else:
            status_fill  = fill(RED_BG)
            status_color = "8B0000"
            link_label   = "—"
            notes        = r.get("error", "")

        row_data = [
            r.get("pipeline_name", ""),
            status,
            r.get("build_number", "—"),
            r.get("branch", CONFIG["dev_branch"]),
            r.get("requested_by", "—"),
            r.get("queued_at", "—"),
            notes,
            link_label,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border()
            cell.alignment = left() if col_idx in (1, 7) else center()
            cell.font = cell_font()

            if col_idx == 2:
                cell.fill = status_fill
                cell.font = Font(name="Arial", size=10, bold=True, color=status_color)
            else:
                cell.fill = fill(bg)

            if col_idx == 8 and status == "Triggered":
                cell.hyperlink = r.get("build_url", "")
                cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")

        ws.row_dimensions[row_idx].height = 18

    ws.freeze_panes = "A7"

    # ── Sheet 2: How To Use ───────────────────────────────────────────────
    ws2 = wb.create_sheet("How To Use")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 90

    instructions = [
        ("DEV Build Creator — How To Use", BLUE_DARK, 14, True),
        ("", WHITE, 10, False),
        ("PURPOSE", BLUE_MID, 11, True),
        ("  Run this script before a QC Release to automatically queue a DEVELOPMENT build", WHITE, 10, False),
        ("  for every pipeline. The generated Excel lets the approver review and approve each build.", WHITE, 10, False),
        ("", WHITE, 10, False),
        ("STEP 1 — Configure the script", BLUE_MID, 11, True),
        ("  Open dev_build_creator.py and fill in the CONFIG section at the top:", WHITE, 10, False),
        ("  • org          → Your Azure DevOps organisation name", WHITE, 10, False),
        ("  • project      → Your project name", WHITE, 10, False),
        ("  • pat          → A PAT with Build: Read AND Build: Execute (Queue) permissions", WHITE, 10, False),
        ("  • dev_branch   → Branch to build (default: 'development')", WHITE, 10, False),
        ("  • pipeline_names → Leave as [] to trigger ALL pipelines, or list specific ones", WHITE, 10, False),
        ("", WHITE, 10, False),
        ("STEP 2 — Create / update your PAT token", BLUE_MID, 11, True),
        ("  1. Go to Azure DevOps → User Settings (top right) → Personal Access Tokens", WHITE, 10, False),
        ("  2. Click 'New Token'", WHITE, 10, False),
        ("  3. Set scope: Build → Read AND Execute (Queue)", WHITE, 10, False),
        ("  4. Copy the token into CONFIG['pat'] above", WHITE, 10, False),
        ("", WHITE, 10, False),
        ("STEP 3 — Run the script", BLUE_MID, 11, True),
        ("  pip install requests openpyxl", WHITE, 10, False),
        ("  python dev_build_creator.py", WHITE, 10, False),
        ("", WHITE, 10, False),
        ("STEP 4 — Share the output with the approver", BLUE_MID, 11, True),
        ("  The script produces: dev_build_report_YYYYMMDD_HHMM.xlsx", WHITE, 10, False),
        ("  The approver can click the 'Open ↗' links to navigate directly to each triggered build", WHITE, 10, False),
        ("  and begin the approval process without having to manually open each pipeline.", WHITE, 10, False),
        ("", WHITE, 10, False),
        ("UNDERSTANDING THE STATUS COLOURS", BLUE_MID, 11, True),
        ("  Triggered        → Build was successfully queued on the dev branch", WHITE, 10, False),
        ("  Failed to Trigger → Build could not be queued (see Notes column for reason)", WHITE, 10, False),
        ("", WHITE, 10, False),
        ("COMMON FAILURE REASONS", BLUE_MID, 11, True),
        ("  Branch not found → The pipeline does not have the configured dev branch", WHITE, 10, False),
        ("  Permission denied → PAT is missing Build: Execute (Queue) permission", WHITE, 10, False),
    ]

    for i, (text, bg_hex, font_size, bold) in enumerate(instructions, start=1):
        cell = ws2.cell(row=i, column=1, value=text)
        cell.font = Font(name="Arial", size=font_size, bold=bold,
                         color=WHITE if bg_hex != WHITE else "000000")
        cell.fill = fill(bg_hex) if bg_hex != WHITE else PatternFill()
        cell.alignment = left()
        ws2.row_dimensions[i].height = 18

    return wb


def main():
    print(f"\n{'='*55}")
    print("  DEV Build Creator — Azure DevOps")
    print(f"{'='*55}")
    print(f"  Org:        {CONFIG['org']}")
    print(f"  Project:    {CONFIG['project']}")
    print(f"  Branch:     {CONFIG['dev_branch']}")
    print(f"{'='*55}\n")

    print("Fetching pipelines...")
    try:
        pipelines = get_all_pipelines()
    except Exception as e:
        print(f"Failed to connect to Azure DevOps: {e}")
        print("   → Check your org name, project name, and PAT token in CONFIG")
        return

    if not pipelines:
        print("No pipelines found. Check your CONFIG filters.")
        return

    print(f"   Found {len(pipelines)} pipeline(s). Triggering builds in parallel...\n")

    def process_pipeline(pipeline):
        print(f"  Queuing: {pipeline['name']}")
        result = trigger_dev_build(pipeline["id"], pipeline["name"])
        result["pipeline_name"] = pipeline["name"]
        return result

    total_pipelines = len(pipelines)
    results = []
    completed = 0
    try:
        with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
            futures = {executor.submit(process_pipeline, p): p for p in pipelines}
            for future in concurrent.futures.as_completed(futures):
                result = future.result()
                results.append(result)
                completed += 1
                status = result["status"]
                name   = result["pipeline_name"]
                prefix = f"  [{completed}/{total_pipelines}]"
                if status == "Triggered":
                    print(f"{prefix} Triggered  {name} — Build #{result['build_number']} | {result['queued_at']}")
                else:
                    print(f"{prefix} FAILED     {name} — {result.get('error', '')}")
    except KeyboardInterrupt:
        print("\n\n  Interrupted by user — exiting.")
        os._exit(1)

    # Sort results: triggered first, then failed — makes the Excel easier to read
    results.sort(key=lambda r: (0 if r["status"] == "Triggered" else 1, r.get("pipeline_name", "")))

    # Generate outputs
    output_dir = CONFIG.get("output_dir", ".")
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
    xlsx_file = os.path.join(output_dir, f"dev_build_report_{timestamp}.xlsx")
    json_file = os.path.join(output_dir, f"dev_build_report_{timestamp}.json")

    print(f"\nBuilding Excel report...")
    wb = build_excel_report(results)
    wb.save(xlsx_file)

    print(f"Saving JSON...")
    with open(json_file, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, default=str)

    triggered = sum(1 for r in results if r["status"] == "Triggered")
    failed    = sum(1 for r in results if r["status"] == "Failed to Trigger")

    print(f"\n{'='*55}")
    print(f"  Reports saved:")
    print(f"       Excel: {xlsx_file}")
    print(f"       JSON:  {json_file}")
    print(f"{'='*55}")
    print(f"  Pipelines scanned  : {len(results)}")
    print(f"  Builds triggered   : {triggered}")
    print(f"  Failed to trigger  : {failed}")
    print(f"{'='*55}\n")

    if failed:
        print("  Pipelines that failed to trigger:")
        for r in results:
            if r["status"] == "Failed to Trigger":
                print(f"    - {r['pipeline_name']}: {r.get('error', '')}")
        print()


if __name__ == "__main__":
    main()
