"""
UAT Build Tracker
-----------------
Scans all your Azure DevOps repos, finds the last build successfully
deployed to QC, and exports a ready-to-share Excel report.

SETUP (one time):
  pip install requests openpyxl

CONFIGURE: Fill in the CONFIG section below, then run:
  python uat_build_tracker.py
"""

import concurrent.futures
import os
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import json
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Reports are written to <project-root>/reports/uat/ regardless of where the script is run from.
_REPORTS_DIR = os.path.normpath(
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "reports", "uat")
)

# ─────────────────────────────────────────────
#  CONFIG  ← Fill these in before running
# ─────────────────────────────────────────────
CONFIG = {
    "org":        "scmdevops",          # Azure DevOps org name
    "project":    "Leo.TPRM",       # Project name
    "pat":        "<YOUR_PAT_HERE>",  # PAT with Read access to Builds
    # Name of the stage/environment in your pipelines that represents QC
    # Common values: "QC", "QA", "Quality", "Test", "Staging"
    "qc_stage_name": "DEV: QC",
    # Stages that count as "higher" than QC — builds already in these will be EXCLUDED
    # Based on your pipeline, these are the UAT and PROD stage keywords
    "higher_env_stages": ["UAT", "PROD"],
    # Leave pipeline_names as [] to auto-discover all pipelines in the project.
    # Or list specific names to scan only those pipelines.
    "pipeline_names": [],
    # Pipelines whose name contains any of these substrings (case-insensitive) are skipped.
    "exclude_pipeline_name_patterns": ["-CI", "plugin", "automation", "NotInUse", "Not_in_use", "Cypress"],
    # Per-pipeline variable checks: builds will be flagged if any variable doesn't match the required value.
    # Keys are pipeline names; values are dicts of { variable_name: required_value }.
    "pipeline_variable_checks": {
        "leo.tprm.camunda.bpmn_New": {"clientname": "ALL"}
    },
    # Builds from branches whose name contains any of these substrings (case-insensitive)
    # will be ignored entirely. RELEASE builds have their own QC/UAT paths and must not
    # interfere with the DEVELOPMENT/HOTFIX scan.
    "exclude_source_branches": ["release"],
    # Directory where Excel and JSON reports are saved. Created automatically if it doesn't exist.
    "output_dir": _REPORTS_DIR,
}
# ─────────────────────────────────────────────

# ── Load shared credentials from config.json at the project root ────────────
_CONFIG_FILE = os.path.normpath(
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "config.json")
)
try:
    with open(_CONFIG_FILE, encoding="utf-8") as _f:
        _shared = json.load(_f)
    CONFIG["org"]     = _shared.get("org",     CONFIG["org"])
    CONFIG["project"] = _shared.get("project", CONFIG["project"])
    CONFIG["pat"]     = _shared.get("pat",      CONFIG["pat"])
except FileNotFoundError:
    pass  # Falls back to the values set in CONFIG above
# ────────────────────────────────────────────────────────────────────────────

BASE_URL = f"https://dev.azure.com/{CONFIG['org']}/{CONFIG['project']}/_apis"
AUTH = ("", CONFIG["pat"])
HEADERS = {"Content-Type": "application/json"}

# Shared session with connection pooling and automatic retries.
# Reusing connections avoids opening a new socket per request, which
# prevents Windows ephemeral-port exhaustion (WinError 10048).
_session = requests.Session()
_session.auth = AUTH
_session.headers.update(HEADERS)
_retry = Retry(total=3, backoff_factor=0.5, status_forcelist=[429, 500, 502, 503, 504])
_adapter = HTTPAdapter(pool_connections=20, pool_maxsize=20, max_retries=_retry)
_session.mount("https://", _adapter)


def get(url, params=None):
    params = params or {}
    params["api-version"] = "7.1"
    # (connect_timeout, read_timeout) — avoids infinite hangs on slow/unresponsive endpoints
    r = _session.get(url, params=params, timeout=(10, 30))
    r.raise_for_status()
    return r.json()


def check_build_variables(build, pipeline_name):
    """
    Check whether a build satisfies the per-pipeline variable constraints in CONFIG.
    Returns (passes: bool, detail: str | None).
    detail is e.g. "clientname='TENANT_X'" when the check fails.
    """
    checks = CONFIG.get("pipeline_variable_checks", {}).get(pipeline_name, {})
    if not checks:
        return True, None

    # templateParameters covers YAML pipeline parameters overridden at queue time.
    template_params = build.get("templateParameters") or {}
    # parameters is a JSON string of queue-time variable overrides.
    try:
        params = json.loads(build.get("parameters") or "{}")
    except (json.JSONDecodeError, TypeError):
        params = {}

    failures = []
    for var_name, expected in checks.items():
        actual = template_params.get(var_name) or params.get(var_name)
        if not actual or str(actual).strip() == "":
            # Variable not explicitly set → pipeline uses its YAML default → treat as valid.
            continue
        if str(actual).strip().upper() != str(expected).strip().upper():
            failures.append(f"{var_name}={actual!r}")

    if failures:
        return False, ", ".join(failures)
    return True, None


def get_all_pipelines():
    """Fetch all pipelines in the project using the /pipelines endpoint.

    The /pipelines endpoint returns Pipeline IDs that are consistent with the
    Azure DevOps UI (definitionId in the URL). These IDs are also accepted by
    /pipelines/{id}/runs which is used to fetch recent builds.
    """
    data = get(f"{BASE_URL}/pipelines", {"$top": 500})
    pipelines = data.get("value", [])

    if CONFIG["pipeline_names"]:
        # Explicit allowlist takes priority — only scan the named pipelines.
        pipelines = [p for p in pipelines if p["name"] in CONFIG["pipeline_names"]]
    else:
        # Auto-discover: exclude pipelines whose name contains any exclusion pattern.
        exclude_patterns = [p.lower() for p in CONFIG.get("exclude_pipeline_name_patterns", [])]
        if exclude_patterns:
            pipelines = [
                p for p in pipelines
                if not any(pat in p["name"].lower() for pat in exclude_patterns)
            ]

    return pipelines


def get_last_qc_build(pipeline_id, pipeline_name):
    """
    For a given pipeline, find the most recent build that passed QC but not UAT.

    Uses /build/builds with separate completed + inProgress queries (with
    queryOrder=queueTimeDescending) to ensure recently-triggered in-progress builds
    are included. Without explicit queryOrder, the API returns oldest-first and
    newer in-progress builds can be missed. Timelines are fetched in parallel.
    """
    try:
        print(f"    [{pipeline_name}] fetching builds...")
        # $top=50: buffer large enough to find dev/hotfix builds even if recent history
        # is dominated by release-branch builds that get filtered out below.
        completed_data = get(f"{BASE_URL}/build/builds", {
            "definitions": pipeline_id,
            "$top": 50,
            "statusFilter": "completed",
            "queryOrder": "queueTimeDescending",
        })
        inprogress_data = get(f"{BASE_URL}/build/builds", {
            "definitions": pipeline_id,
            "$top": 50,
            "statusFilter": "inProgress",
            "queryOrder": "queueTimeDescending",
        })
        seen = set()
        raw = []
        for b in inprogress_data.get("value", []) + completed_data.get("value", []):
            if b["id"] not in seen:
                seen.add(b["id"])
                raw.append(b)
        builds = sorted(raw, key=lambda b: b["id"], reverse=True)
        print(f"    [{pipeline_name}] fetched {len(builds)} builds ({len(inprogress_data.get('value',[]))} inProgress + {len(completed_data.get('value',[]))} completed)")
    except Exception as e:
        print(f"    [{pipeline_name}] ERROR fetching builds: {e}")
        return {"status": "API Error", "error": str(e)}

    if not builds:
        return {"status": "No QC Build Found"}

    # Filter out excluded branches (e.g. release)
    excluded = CONFIG.get("exclude_source_branches", [])
    eligible = [
        b for b in builds
        if not (excluded and any(
            excl.lower() in b.get("sourceBranch", "").lower()
            for excl in excluded
        ))
    ]

    if not eligible:
        return {"status": "No QC Build Found"}

    # Keyed by build id — same structure used by the rest of the function
    build_map = {b["id"]: b for b in eligible}

    print(f"    [{pipeline_name}] {len(eligible)} eligible builds — analysing in batches...")

    def _fetch_timeline(run_id):
        try:
            return run_id, get(f"{BASE_URL}/build/builds/{run_id}/timeline").get("records", [])
        except Exception:
            return run_id, []

    # Fetch timelines in small batches (newest-first) and stop as soon as the
    # answer is known — avoids fetching all ~50+ timelines when the qualifying
    # build is typically one of the first few.
    BATCH_SIZE = 3
    found_uat_newer = False
    uat_superseding = {}

    for i in range(0, len(eligible), BATCH_SIZE):
        batch = eligible[i:i + BATCH_SIZE]
        with concurrent.futures.ThreadPoolExecutor(max_workers=len(batch)) as ex:
            batch_timelines = dict(ex.map(lambda r: _fetch_timeline(r["id"]), batch))

        for run in batch:
            build_id = run["id"]
            build    = build_map[build_id]
            records  = batch_timelines.get(build_id, [])

            qc_stage = next(
                (r for r in records
                 if CONFIG["qc_stage_name"].lower() in r.get("name", "").lower()
                 and r.get("result") == "succeeded"),
                None,
            )
            higher_env_reached = any(
                any(kw.lower() in r.get("name", "").lower() for kw in CONFIG["higher_env_stages"])
                and r.get("result") in ("succeeded", "partiallySucceeded")
                for r in records
            )

            if higher_env_reached and not found_uat_newer:
                found_uat_newer = True
                uat_superseding = {
                    "uat_build_number": build.get("buildNumber", "N/A"),
                    "uat_build_url": f"https://{CONFIG['org']}.visualstudio.com/{CONFIG['project']}/_build/results?buildId={build_id}",
                }

            if qc_stage and not higher_env_reached:
                if found_uat_newer:
                    return {"status": "Superseded by UAT", **uat_superseding}

                finished = build.get("finishTime", "")
                finished_dt = datetime.strptime(finished[:19], "%Y-%m-%dT%H:%M:%S") if finished else None
                build_info = {
                    "build_id": build_id,
                    "build_number": build.get("buildNumber", "N/A"),
                    "branch": build.get("sourceBranch", "").replace("refs/heads/", ""),
                    "commit": build.get("sourceVersion", "")[:8],
                    "requested_by": build.get("requestedFor", {}).get("displayName", "N/A"),
                    "qc_deployed_at": finished_dt.strftime("%Y-%m-%d %H:%M UTC") if finished_dt else "N/A",
                    "build_url": f"https://{CONFIG['org']}.visualstudio.com/{CONFIG['project']}/_build/results?buildId={build_id}",
                }
                passes_vars, var_detail = check_build_variables(build, pipeline_name)
                if not passes_vars:
                    return {"status": "Wrong Variables", "variable_detail": var_detail, **build_info}
                return {"status": "Found", **build_info}

    return {"status": "No QC Build Found"}


def build_excel_report(results):
    wb = Workbook()

    # ── Styles ──────────────────────────────────────────────────────────────
    BLUE_DARK   = "1F3864"
    BLUE_MID    = "2E75B6"
    BLUE_LIGHT  = "D6E4F0"
    GREEN_BG    = "E2EFDA"
    AMBER_BG    = "FFF2CC"
    RED_BG      = "FFDDD8"
    WHITE       = "FFFFFF"
    GREY_ROW    = "F5F7FA"

    def header_font(size=11, bold=True, color=WHITE):
        return Font(name="Arial", size=size, bold=bold, color=color)

    def cell_font(size=10, bold=False, color="000000"):
        return Font(name="Arial", size=size, bold=bold, color=color)

    def fill(hex_color):
        return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

    def thin_border():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def left():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ── Sheet 1: Summary Dashboard ────────────────────────────────────────
    ws = wb.active
    ws.title = "UAT Release Report"
    ws.sheet_view.showGridLines = False

    # Title block
    ws.merge_cells("A1:H1")
    ws["A1"] = "UAT Release — QC Build Report"
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color=WHITE)
    ws["A1"].fill = fill(BLUE_DARK)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}  |  Org: {CONFIG['org']}  |  Project: {CONFIG['project']}"
    ws["A2"].font = Font(name="Arial", size=9, color=WHITE)
    ws["A2"].fill = fill(BLUE_MID)
    ws["A2"].alignment = center()
    ws.row_dimensions[2].height = 18

    # Summary KPI row
    total      = len(results)
    found      = sum(1 for r in results if r["status"] == "Found")
    not_found  = sum(1 for r in results if r["status"] == "No QC Build Found")
    errors     = sum(1 for r in results if r["status"] == "API Error")

    ws.merge_cells("A3:B3"); ws["A3"] = "Total Repos Scanned"
    ws.merge_cells("C3:D3"); ws["C3"] = "QC Build Found"
    ws.merge_cells("E3:F3"); ws["E3"] = "No QC Build"
    ws.merge_cells("G3:H3"); ws["G3"] = "Errors"
    ws.merge_cells("A4:B4"); ws["A4"] = total
    ws.merge_cells("C4:D4"); ws["C4"] = found
    ws.merge_cells("E4:F4"); ws["E4"] = not_found
    ws.merge_cells("G4:H4"); ws["G4"] = errors

    for col_letter, bg in [("A", BLUE_LIGHT), ("C", GREEN_BG), ("E", AMBER_BG), ("G", RED_BG)]:
        for row in [3, 4]:
            cell = ws[f"{col_letter}{row}"]
            cell.fill = fill(bg)
            cell.alignment = center()
            if row == 3:
                cell.font = cell_font(9, bold=True)
            else:
                cell.font = Font(name="Arial", size=20, bold=True)

    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 40

    # Blank spacer
    ws.row_dimensions[5].height = 8

    # Column headers
    headers = ["Repo / Pipeline", "Status", "Build #", "Branch", "Commit", "Deployed By", "QC Deployed At", "Build Link"]
    col_widths = [35, 18, 14, 22, 12, 24, 22, 14]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=6, column=col_idx, value=h)
        cell.font = header_font()
        cell.fill = fill(BLUE_MID)
        cell.alignment = center()
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[6].height = 22

    # Data rows
    for row_idx, r in enumerate(results, start=7):
        bg = GREY_ROW if row_idx % 2 == 0 else WHITE
        status = r["status"]

        if status == "Found":
            status_fill = fill(GREEN_BG)
            status_color = "276221"
        elif status == "No QC Build Found":
            status_fill = fill(AMBER_BG)
            status_color = "7D5200"
        elif status == "Superseded by UAT":
            status_fill = fill("FDE9D9")
            status_color = "843C0C"
        elif status == "Wrong Variables":
            status_fill = fill("EAD1F5")
            status_color = "5B0D91"
        else:
            status_fill = fill(RED_BG)
            status_color = "8B0000"

        # For Wrong Variables, show the actual variable values in the status cell text.
        display_status = status
        if status == "Wrong Variables" and r.get("variable_detail"):
            display_status = f"Wrong Variables\n({r['variable_detail']})"

        if status == "Superseded by UAT":
            build_num_display = f"UAT #{r.get('uat_build_number', '?')}"
            link_label = "Open UAT ↗"
        else:
            build_num_display = r.get("build_number", "—")
            link_label = "Open ↗" if status in ("Found", "Wrong Variables") else "—"

        row_data = [
            r.get("pipeline_name", ""),
            display_status,
            build_num_display,
            r.get("branch", "—"),
            r.get("commit", "—"),
            r.get("requested_by", "—"),
            r.get("qc_deployed_at", "—"),
            link_label,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border()
            cell.alignment = left() if col_idx == 1 else center()
            cell.font = cell_font()

            if col_idx == 2:
                cell.fill = status_fill
                cell.font = Font(name="Arial", size=10, bold=True, color=status_color)
            else:
                cell.fill = fill(bg)

            # Hyperlink for Build Link column
            if col_idx == 8 and status in ("Found", "Wrong Variables"):
                cell.hyperlink = r.get("build_url", "")
                cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")
            elif col_idx == 8 and status == "Superseded by UAT" and r.get("uat_build_url"):
                cell.hyperlink = r.get("uat_build_url", "")
                cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")

        ws.row_dimensions[row_idx].height = 18

    # Freeze panes below header
    ws.freeze_panes = "A7"

    # ── Sheet 2: How To Use ───────────────────────────────────────────────
    ws2 = wb.create_sheet("How To Use")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 80

    instructions = [
        ("UAT Build Tracker — How To Use", BLUE_DARK, 14, True),
        ("", WHITE, 10, False),
        ("STEP 1 — Configure the script", BLUE_MID, 11, True),
        ("Open uat_build_tracker.py and fill in the CONFIG section at the top:", WHITE, 10, False),
        ("  • org           → Your Azure DevOps organisation name", WHITE, 10, False),
        ("  • project       → Your project name", WHITE, 10, False),
        ("  • pat           → A Personal Access Token (PAT) with Build: Read permission", WHITE, 10, False),
        ("  • qc_stage_name → The exact name of your QC stage in pipelines (e.g. 'QC', 'QA')", WHITE, 10, False),
        ("  • pipeline_names → Leave as [] to scan ALL pipelines, or list specific ones", WHITE, 10, False),
        ("", WHITE, 10, False),
        ("STEP 2 — Create a PAT token (if you don't have one)", BLUE_MID, 11, True),
        ("  1. Go to Azure DevOps → User Settings (top right) → Personal Access Tokens", WHITE, 10, False),
        ("  2. Click 'New Token'", WHITE, 10, False),
        ("  3. Set scope: Build → Read", WHITE, 10, False),
        ("  4. Copy the token into CONFIG['pat'] above", WHITE, 10, False),
        ("", WHITE, 10, False),
        ("STEP 3 — Run the script", BLUE_MID, 11, True),
        ("  pip install requests openpyxl", WHITE, 10, False),
        ("  python uat_build_tracker.py", WHITE, 10, False),
        ("", WHITE, 10, False),
        ("STEP 4 — Share the output", BLUE_MID, 11, True),
        ("  The script produces: uat_release_report_YYYYMMDD_HHMM.xlsx", WHITE, 10, False),
        ("  Share this file with your DevOps team before each UAT release.", WHITE, 10, False),
        ("  They can click the 'Open ↗' links to jump directly to the right build.", WHITE, 10, False),
        ("", WHITE, 10, False),
        ("UNDERSTANDING THE STATUS COLOURS", BLUE_MID, 11, True),
        ("  ✅ Found              → A successful QC deployment was identified", WHITE, 10, False),
        ("  ⚠️  No QC Build Found → Pipeline exists but no completed QC stage found in last 20 builds", WHITE, 10, False),
        ("  ❌ API Error          → Could not reach the pipeline (check PAT permissions)", WHITE, 10, False),
        ("", WHITE, 10, False),
        ("RUNNING ON A SCHEDULE", BLUE_MID, 11, True),
        ("  Windows: Use Task Scheduler to run this script daily before standups.", WHITE, 10, False),
        ("  Mac/Linux: Add a cron job:  0 8 * * * python /path/to/uat_build_tracker.py", WHITE, 10, False),
    ]

    for i, (text, bg_hex, font_size, bold) in enumerate(instructions, start=1):
        cell = ws2.cell(row=i, column=1, value=text)
        cell.font = Font(name="Arial", size=font_size, bold=bold,
                         color=WHITE if bg_hex != WHITE else "000000")
        cell.fill = fill(bg_hex) if bg_hex != WHITE else PatternFill()
        cell.alignment = left()
        ws2.row_dimensions[i].height = 18

    return wb


def main():
    print(f"\n{'='*55}")
    print("  UAT Build Tracker — Azure DevOps")
    print(f"{'='*55}")
    print(f"  Org:     {CONFIG['org']}")
    print(f"  Project: {CONFIG['project']}")
    print(f"  QC Stage: '{CONFIG['qc_stage_name']}'")
    print(f"{'='*55}\n")

    print("📋 Fetching pipelines...")
    try:
        pipelines = get_all_pipelines()
    except Exception as e:
        print(f"❌ Failed to connect to Azure DevOps: {e}")
        print("   → Check your org name, project name, and PAT token in CONFIG")
        return

    print(f"   Scanning {len(pipelines)} pipeline(s) in parallel...\n")

    def scan_pipeline(pipeline):
        print(f"  🔍 Processing: {pipeline['name']}")
        result = get_last_qc_build(pipeline["id"], pipeline["name"])
        result["pipeline_name"] = pipeline["name"]
        return result

    total_pipelines = len(pipelines)
    results = []
    completed = 0
    try:
        with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
            futures = {executor.submit(scan_pipeline, p): p for p in pipelines}
            for future in concurrent.futures.as_completed(futures):
                result = future.result()
                results.append(result)
                completed += 1
                status = result["status"]
                name   = result["pipeline_name"]
                prefix = f"  [{completed}/{total_pipelines}]"
                if status == "Found":
                    print(f"{prefix} ✅  {name} — Build #{result['build_number']} | {result['qc_deployed_at']}")
                elif status == "Superseded by UAT":
                    print(f"{prefix} 🔁  {name} — Superseded by UAT Build #{result.get('uat_build_number', '?')}")
                elif status == "Wrong Variables":
                    print(f"{prefix} 🚫  {name} — Wrong variables ({result.get('variable_detail', '')})")
                elif status == "No QC Build Found":
                    print(f"{prefix} ⚠️   {name} — No QC build found")
                else:
                    print(f"{prefix} ❌  {name} — Error: {result.get('error', '')}")
    except KeyboardInterrupt:
        print("\n\n  Interrupted by user — exiting.")
        os._exit(1)

    # Generate outputs
    output_dir = CONFIG.get("output_dir", ".")
    os.makedirs(output_dir, exist_ok=True)
    timestamp   = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
    xlsx_file   = os.path.join(output_dir, f"uat_release_report_{timestamp}.xlsx")
    json_file   = os.path.join(output_dir, f"uat_release_report_{timestamp}.json")

    print(f"\n📊 Building Excel report...")
    wb = build_excel_report(results)
    wb.save(xlsx_file)

    print(f"💾 Saving JSON...")
    with open(json_file, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, default=str)

    found       = sum(1 for r in results if r["status"] == "Found")
    not_found   = sum(1 for r in results if r["status"] == "No QC Build Found")
    superseded  = sum(1 for r in results if r["status"] == "Superseded by UAT")
    wrong_vars  = sum(1 for r in results if r["status"] == "Wrong Variables")
    errors      = sum(1 for r in results if r["status"] == "API Error")

    print(f"\n{'='*55}")
    print(f"  ✅  Reports saved:")
    print(f"       Excel: {xlsx_file}")
    print(f"       JSON:  {json_file}")
    print(f"{'='*55}")
    print(f"  Repos scanned    : {len(results)}")
    print(f"  QC build found   : {found}")
    print(f"  Superseded by UAT: {superseded}")
    print(f"  Wrong variables  : {wrong_vars}")
    print(f"  Not found        : {not_found}")
    print(f"  Errors           : {errors}")
    print(f"{'='*55}\n")


if __name__ == "__main__":
    main()
