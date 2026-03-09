"""
PROD Build Tracker
------------------
Scans all your Azure DevOps repos, finds the last build successfully
deployed to UAT (and not yet to PROD), and exports a ready-to-share Excel report.

SETUP (one time):
  pip install requests openpyxl

CONFIGURE: Fill in the CONFIG section below, then run:
  python prod_build_tracker.py
"""

import concurrent.futures
import json
import os
import sys
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import common

# Reports are written to <project-root>/reports/prod/ regardless of where the script is run from.
_REPORTS_DIR = os.path.normpath(
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "reports", "prod")
)

# ─────────────────────────────────────────────
#  CONFIG  ← Fill these in before running
# ─────────────────────────────────────────────
CONFIG = {
    "org":        "scmdevops",          # Azure DevOps org name
    "project":    "Leo.TPRM",           # Project name
    "pat":        "<YOUR_PAT_HERE>",  # PAT with Read access to Builds
    # Name(s) of the stage in your pipelines that represent a successful UAT deployment.
    # Can be a single string or a list of strings — a build qualifies if ANY of these
    # stage names is found (case-insensitive substring match).
    # Common values: "Deploy to UAT-US", "REL: UAT Deployment"
    "uat_stage_name": ["Deploy to UAT-US", "REL: UAT Deployment"],
    # Stages that count as "higher" than UAT — builds already in these will be EXCLUDED
    "higher_env_stages": ["PROD"],
    # Leave pipeline_names as [] to auto-discover all pipelines in the project.
    # Or list specific names to scan only those pipelines.
    "pipeline_names": [],
    # Pipelines whose name contains any of these substrings (case-insensitive) are skipped.
    "exclude_pipeline_name_patterns": ["-CI", "plugin", "automation", "NotInUse", "Not_in_use", "Cypress"],
    # Per-pipeline variable checks: builds will be flagged if any variable doesn't match the required value.
    # Keys are pipeline names; values are dicts of { variable_name: required_value }.
    "pipeline_variable_checks": {},
    # Builds from branches whose name contains any of these substrings (case-insensitive)
    # will be ignored entirely. Leave as [] to include all branches (including RELEASE).
    "exclude_source_branches": [],
    # Builds from branches matching these patterns (case-insensitive) that have reached PROD
    # will NOT be treated as superseding earlier UAT builds from RELEASE/DEVELOPMENT branches.
    # A hotfix PROD deployment is a bug fix — it should not block surfacing the next release candidate.
    "hotfix_branch_patterns": ["hotfix"],
    # Directory where Excel and JSON reports are saved. Created automatically if it doesn't exist.
    "output_dir": _REPORTS_DIR,
}
# ─────────────────────────────────────────────

common.load_credentials(__file__, CONFIG)
_session, BASE_URL = common.build_session(CONFIG)
get, post, patch = common.make_http_fns(_session)


def check_build_variables(build, pipeline_name):
    """
    Check whether a build satisfies the per-pipeline variable constraints in CONFIG.
    Returns (passes: bool, detail: str | None).
    detail is e.g. "clientname='TENANT_X'" when the check fails.
    """
    checks = CONFIG.get("pipeline_variable_checks", {}).get(pipeline_name, {})
    if not checks:
        return True, None

    template_params = build.get("templateParameters") or {}
    try:
        params = json.loads(build.get("parameters") or "{}")
    except (json.JSONDecodeError, TypeError):
        params = {}

    failures = []
    for var_name, expected in checks.items():
        actual = template_params.get(var_name) or params.get(var_name)
        if not actual or str(actual).strip() == "":
            continue
        if str(actual).strip().upper() != str(expected).strip().upper():
            failures.append(f"{var_name}={actual!r}")

    if failures:
        return False, ", ".join(failures)
    return True, None


def get_all_pipelines():
    """Fetch all pipelines in the project using the /pipelines endpoint."""
    data = get(f"{BASE_URL}/pipelines", {"$top": 500})
    pipelines = data.get("value", [])

    if CONFIG["pipeline_names"]:
        pipelines = [p for p in pipelines if p["name"] in CONFIG["pipeline_names"]]
    else:
        exclude_patterns = [p.lower() for p in CONFIG.get("exclude_pipeline_name_patterns", [])]
        if exclude_patterns:
            pipelines = [
                p for p in pipelines
                if not any(pat in p["name"].lower() for pat in exclude_patterns)
            ]

    return pipelines


def get_last_uat_build(pipeline_id, pipeline_name):
    """
    For a given pipeline, find the most recent build that passed UAT but not PROD.

    Uses /build/builds with separate completed + inProgress queries (with
    queryOrder=queueTimeDescending) to ensure recently-triggered in-progress builds
    are included. Without explicit queryOrder, the API returns oldest-first and
    newer in-progress builds can be missed. Timelines are fetched in parallel.
    """
    try:
        print(f"    [{pipeline_name}] fetching builds...")
        completed_data = get(f"{BASE_URL}/build/builds", {
            "definitions": pipeline_id,
            "$top": 50,
            "statusFilter": "completed",
            "queryOrder": "queueTimeDescending",
        })
        inprogress_data = get(f"{BASE_URL}/build/builds", {
            "definitions": pipeline_id,
            "$top": 50,
            "statusFilter": "inProgress",
            "queryOrder": "queueTimeDescending",
        })
        seen = set()
        raw = []
        for b in inprogress_data.get("value", []) + completed_data.get("value", []):
            if b["id"] not in seen:
                seen.add(b["id"])
                raw.append(b)
        builds = sorted(raw, key=lambda b: b["id"], reverse=True)
        print(f"    [{pipeline_name}] fetched {len(builds)} builds ({len(inprogress_data.get('value',[]))} inProgress + {len(completed_data.get('value',[]))} completed)")
    except Exception as e:
        print(f"    [{pipeline_name}] ERROR fetching builds: {e}")
        return {"status": "API Error", "error": str(e)}

    if not builds:
        return {"status": "No UAT Build Found"}

    # Filter out excluded branches
    excluded = CONFIG.get("exclude_source_branches", [])
    eligible = [
        b for b in builds
        if not (excluded and any(
            excl.lower() in b.get("sourceBranch", "").lower()
            for excl in excluded
        ))
    ]

    if not eligible:
        return {"status": "No UAT Build Found"}

    build_map = {b["id"]: b for b in eligible}

    print(f"    [{pipeline_name}] {len(eligible)} eligible builds — fetching timelines...")

    def _fetch_timeline(run_id):
        try:
            return run_id, get(f"{BASE_URL}/build/builds/{run_id}/timeline").get("records", [])
        except Exception:
            return run_id, []

    with concurrent.futures.ThreadPoolExecutor(max_workers=min(len(eligible), 5)) as ex:
        timeline_map = dict(ex.map(lambda r: _fetch_timeline(r["id"]), eligible))

    print(f"    [{pipeline_name}] timelines fetched — analysing...")

    # Process newest → oldest
    found_prod_newer = False
    prod_superseding = {}
    hotfix_patterns = [p.lower() for p in CONFIG.get("hotfix_branch_patterns", [])]

    for run in eligible:
        build_id = run["id"]
        build    = build_map[build_id]
        records  = timeline_map.get(build_id, [])

        uat_names = CONFIG["uat_stage_name"]
        if isinstance(uat_names, str):
            uat_names = [uat_names]
        uat_stage = next(
            (r for r in records
             if any(n.lower() in r.get("name", "").lower() for n in uat_names)
             and r.get("result") == "succeeded"),
            None,
        )
        higher_env_reached = any(
            any(kw.lower() in r.get("name", "").lower() for kw in CONFIG["higher_env_stages"])
            and r.get("result") in ("succeeded", "partiallySucceeded")
            for r in records
        )

        is_hotfix_branch = bool(hotfix_patterns and any(
            p in run.get("sourceBranch", "").lower() for p in hotfix_patterns
        ))

        if higher_env_reached and not found_prod_newer and not is_hotfix_branch:
            found_prod_newer = True
            prod_superseding = {
                "prod_build_number": build.get("buildNumber", "N/A"),
                "prod_build_url": f"https://{CONFIG['org']}.visualstudio.com/{CONFIG['project']}/_build/results?buildId={build_id}",
            }

        if uat_stage and not higher_env_reached:
            if found_prod_newer:
                return {"status": "Superseded by PROD", **prod_superseding}

            finished = build.get("finishTime", "")
            finished_dt = datetime.strptime(finished[:19], "%Y-%m-%dT%H:%M:%S") if finished else None
            build_info = {
                "build_id": build_id,
                "build_number": build.get("buildNumber", "N/A"),
                "branch": build.get("sourceBranch", "").replace("refs/heads/", ""),
                "commit": build.get("sourceVersion", "")[:8],
                "requested_by": build.get("requestedFor", {}).get("displayName", "N/A"),
                "uat_deployed_at": finished_dt.strftime("%Y-%m-%d %H:%M UTC") if finished_dt else "N/A",
                "build_url": f"https://{CONFIG['org']}.visualstudio.com/{CONFIG['project']}/_build/results?buildId={build_id}",
            }
            passes_vars, var_detail = check_build_variables(build, pipeline_name)
            if not passes_vars:
                return {"status": "Wrong Variables", "variable_detail": var_detail, **build_info}
            return {"status": "Found", **build_info}

    return {"status": "No UAT Build Found"}


def build_excel_report(results):
    wb = Workbook()

    from common import (BLUE_DARK, BLUE_MID, BLUE_LIGHT, GREEN_BG, AMBER_BG, RED_BG,
                        WHITE, GREY_ROW, header_font, cell_font, fill, thin_border, center, left)

    # ── Sheet 1: Summary Dashboard ────────────────────────────────────────
    ws = wb.active
    ws.title = "PROD Release Report"
    ws.sheet_view.showGridLines = False

    # Title block
    ws.merge_cells("A1:H1")
    ws["A1"] = "PROD Release — UAT Build Report"
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color=WHITE)
    ws["A1"].fill = fill(BLUE_DARK)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}  |  Org: {CONFIG['org']}  |  Project: {CONFIG['project']}"
    ws["A2"].font = Font(name="Arial", size=9, color=WHITE)
    ws["A2"].fill = fill(BLUE_MID)
    ws["A2"].alignment = center()
    ws.row_dimensions[2].height = 18

    # Summary KPI row
    total      = len(results)
    found      = sum(1 for r in results if r["status"] == "Found")
    not_found  = sum(1 for r in results if r["status"] == "No UAT Build Found")
    errors     = sum(1 for r in results if r["status"] == "API Error")

    ws.merge_cells("A3:B3"); ws["A3"] = "Total Repos Scanned"
    ws.merge_cells("C3:D3"); ws["C3"] = "UAT Build Found"
    ws.merge_cells("E3:F3"); ws["E3"] = "No UAT Build"
    ws.merge_cells("G3:H3"); ws["G3"] = "Errors"
    ws.merge_cells("A4:B4"); ws["A4"] = total
    ws.merge_cells("C4:D4"); ws["C4"] = found
    ws.merge_cells("E4:F4"); ws["E4"] = not_found
    ws.merge_cells("G4:H4"); ws["G4"] = errors

    for col_letter, bg in [("A", BLUE_LIGHT), ("C", GREEN_BG), ("E", AMBER_BG), ("G", RED_BG)]:
        for row in [3, 4]:
            cell = ws[f"{col_letter}{row}"]
            cell.fill = fill(bg)
            cell.alignment = center()
            if row == 3:
                cell.font = cell_font(9, bold=True)
            else:
                cell.font = Font(name="Arial", size=20, bold=True)

    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 40

    # Blank spacer
    ws.row_dimensions[5].height = 8

    # Column headers
    headers = ["Repo / Pipeline", "Status", "Build #", "Branch", "Commit", "Deployed By", "UAT Deployed At", "Build Link"]
    col_widths = [35, 18, 14, 22, 12, 24, 22, 14]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=6, column=col_idx, value=h)
        cell.font = header_font()
        cell.fill = fill(BLUE_MID)
        cell.alignment = center()
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[6].height = 22

    # Data rows
    for row_idx, r in enumerate(results, start=7):
        bg = GREY_ROW if row_idx % 2 == 0 else WHITE
        status = r["status"]

        if status == "Found":
            status_fill = fill(GREEN_BG)
            status_color = "276221"
        elif status == "No UAT Build Found":
            status_fill = fill(AMBER_BG)
            status_color = "7D5200"
        elif status == "Superseded by PROD":
            status_fill = fill("FDE9D9")
            status_color = "843C0C"
        elif status == "Wrong Variables":
            status_fill = fill("EAD1F5")
            status_color = "5B0D91"
        else:
            status_fill = fill(RED_BG)
            status_color = "8B0000"

        display_status = status
        if status == "Wrong Variables" and r.get("variable_detail"):
            display_status = f"Wrong Variables\n({r['variable_detail']})"

        if status == "Superseded by PROD":
            build_num_display = f"PROD #{r.get('prod_build_number', '?')}"
            link_label = "Open PROD ↗"
        else:
            build_num_display = r.get("build_number", "—")
            link_label = "Open ↗" if status in ("Found", "Wrong Variables") else "—"

        row_data = [
            r.get("pipeline_name", ""),
            display_status,
            build_num_display,
            r.get("branch", "—"),
            r.get("commit", "—"),
            r.get("requested_by", "—"),
            r.get("uat_deployed_at", "—"),
            link_label,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border()
            cell.alignment = left() if col_idx == 1 else center()
            cell.font = cell_font()

            if col_idx == 2:
                cell.fill = status_fill
                cell.font = Font(name="Arial", size=10, bold=True, color=status_color)
            else:
                cell.fill = fill(bg)

            if col_idx == 8 and status in ("Found", "Wrong Variables"):
                cell.hyperlink = r.get("build_url", "")
                cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")
            elif col_idx == 8 and status == "Superseded by PROD" and r.get("prod_build_url"):
                cell.hyperlink = r.get("prod_build_url", "")
                cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")

        ws.row_dimensions[row_idx].height = 18

    ws.freeze_panes = "A7"

    # ── Sheet 2: How To Use ───────────────────────────────────────────────
    ws2 = wb.create_sheet("How To Use")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 80

    instructions = [
        ("PROD Build Tracker — How To Use", BLUE_DARK, 14, True),
        ("", WHITE, 10, False),
        ("STEP 1 — Configure the script", BLUE_MID, 11, True),
        ("Open prod_build_tracker.py and fill in the CONFIG section at the top:", WHITE, 10, False),
        ("  • org            → Your Azure DevOps organisation name", WHITE, 10, False),
        ("  • project        → Your project name", WHITE, 10, False),
        ("  • pat            → A Personal Access Token (PAT) with Build: Read permission", WHITE, 10, False),
        ("  • uat_stage_name → The exact name of your UAT stage in pipelines (e.g. 'Deploy to UAT')", WHITE, 10, False),
        ("  • pipeline_names → Leave as [] to scan ALL pipelines, or list specific ones", WHITE, 10, False),
        ("", WHITE, 10, False),
        ("STEP 2 — Create a PAT token (if you don't have one)", BLUE_MID, 11, True),
        ("  1. Go to Azure DevOps → User Settings (top right) → Personal Access Tokens", WHITE, 10, False),
        ("  2. Click 'New Token'", WHITE, 10, False),
        ("  3. Set scope: Build → Read", WHITE, 10, False),
        ("  4. Copy the token into CONFIG['pat'] above", WHITE, 10, False),
        ("", WHITE, 10, False),
        ("STEP 3 — Run the script", BLUE_MID, 11, True),
        ("  pip install requests openpyxl", WHITE, 10, False),
        ("  python prod_build_tracker.py", WHITE, 10, False),
        ("", WHITE, 10, False),
        ("STEP 4 — Share the output", BLUE_MID, 11, True),
        ("  The script produces: prod_release_report_YYYYMMDD_HHMM.xlsx", WHITE, 10, False),
        ("  Share this file with your DevOps team before each PROD release.", WHITE, 10, False),
        ("  They can click the 'Open ↗' links to jump directly to the right build.", WHITE, 10, False),
        ("", WHITE, 10, False),
        ("UNDERSTANDING THE STATUS COLOURS", BLUE_MID, 11, True),
        ("  ✅ Found               → A successful UAT deployment was identified, PROD not yet reached", WHITE, 10, False),
        ("  ⚠️  No UAT Build Found → Pipeline exists but no completed UAT stage found in recent builds", WHITE, 10, False),
        ("  🔁 Superseded by PROD  → A newer build has already been deployed to PROD", WHITE, 10, False),
        ("  ❌ API Error           → Could not reach the pipeline (check PAT permissions)", WHITE, 10, False),
        ("", WHITE, 10, False),
        ("RUNNING ON A SCHEDULE", BLUE_MID, 11, True),
        ("  Windows: Use Task Scheduler to run this script daily before standups.", WHITE, 10, False),
        ("  Mac/Linux: Add a cron job:  0 8 * * * python /path/to/prod_build_tracker.py", WHITE, 10, False),
    ]

    for i, (text, bg_hex, font_size, bold) in enumerate(instructions, start=1):
        cell = ws2.cell(row=i, column=1, value=text)
        cell.font = Font(name="Arial", size=font_size, bold=bold,
                         color=WHITE if bg_hex != WHITE else "000000")
        cell.fill = fill(bg_hex) if bg_hex != WHITE else PatternFill()
        cell.alignment = left()
        ws2.row_dimensions[i].height = 18

    return wb


def main():
    print(f"\n{'='*55}")
    print("  PROD Build Tracker — Azure DevOps")
    print(f"{'='*55}")
    print(f"  Org:       {CONFIG['org']}")
    print(f"  Project:   {CONFIG['project']}")
    uat_label = CONFIG['uat_stage_name'] if isinstance(CONFIG['uat_stage_name'], str) else " | ".join(CONFIG['uat_stage_name'])
    print(f"  UAT Stage: '{uat_label}'")
    print(f"{'='*55}\n")

    print("📋 Fetching pipelines...")
    try:
        pipelines = get_all_pipelines()
    except Exception as e:
        print(f"❌ Failed to connect to Azure DevOps: {e}")
        print("   → Check your org name, project name, and PAT token in CONFIG")
        return

    print(f"   Scanning {len(pipelines)} pipeline(s) in parallel...\n")

    def scan_pipeline(pipeline):
        print(f"  🔍 Processing: {pipeline['name']}")
        result = get_last_uat_build(pipeline["id"], pipeline["name"])
        result["pipeline_name"] = pipeline["name"]
        return result

    total_pipelines = len(pipelines)
    results = []
    completed = 0
    try:
        with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
            futures = {executor.submit(scan_pipeline, p): p for p in pipelines}
            for future in concurrent.futures.as_completed(futures):
                result = future.result()
                results.append(result)
                completed += 1
                status = result["status"]
                name   = result["pipeline_name"]
                prefix = f"  [{completed}/{total_pipelines}]"
                if status == "Found":
                    print(f"{prefix} ✅  {name} — Build #{result['build_number']} | {result['uat_deployed_at']}")
                elif status == "Superseded by PROD":
                    print(f"{prefix} 🔁  {name} — Superseded by PROD Build #{result.get('prod_build_number', '?')}")
                elif status == "Wrong Variables":
                    print(f"{prefix} 🚫  {name} — Wrong variables ({result.get('variable_detail', '')})")
                elif status == "No UAT Build Found":
                    print(f"{prefix} ⚠️   {name} — No UAT build found")
                else:
                    print(f"{prefix} ❌  {name} — Error: {result.get('error', '')}")
    except KeyboardInterrupt:
        print("\n\n  Interrupted by user — exiting.")
        os._exit(1)

    # Generate outputs
    output_dir = CONFIG.get("output_dir", ".")
    os.makedirs(output_dir, exist_ok=True)
    timestamp   = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
    xlsx_file   = os.path.join(output_dir, f"prod_release_report_{timestamp}.xlsx")
    json_file   = os.path.join(output_dir, f"prod_release_report_{timestamp}.json")

    print(f"\n📊 Building Excel report...")
    wb = build_excel_report(results)
    wb.save(xlsx_file)

    print(f"💾 Saving JSON...")
    with open(json_file, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, default=str)

    found      = sum(1 for r in results if r["status"] == "Found")
    not_found  = sum(1 for r in results if r["status"] == "No UAT Build Found")
    superseded = sum(1 for r in results if r["status"] == "Superseded by PROD")
    wrong_vars = sum(1 for r in results if r["status"] == "Wrong Variables")
    errors     = sum(1 for r in results if r["status"] == "API Error")

    print(f"\n{'='*55}")
    print(f"  ✅  Reports saved:")
    print(f"       Excel: {xlsx_file}")
    print(f"       JSON:  {json_file}")
    print(f"{'='*55}")
    print(f"  Repos scanned     : {len(results)}")
    print(f"  UAT build found   : {found}")
    print(f"  Superseded by PROD: {superseded}")
    print(f"  Wrong variables   : {wrong_vars}")
    print(f"  Not found         : {not_found}")
    print(f"  Errors            : {errors}")
    print(f"{'='*55}\n")


if __name__ == "__main__":
    main()
