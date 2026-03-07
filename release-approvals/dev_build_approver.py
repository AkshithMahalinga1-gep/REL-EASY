"""
DEV Build Approver
------------------
Reads the latest dev_build_report_*.json, polls each triggered build until it
reaches the "DEV: QC Approval" gate, then auto-approves it via the Azure DevOps
REST API. Exports an Excel + JSON approval report.

SETUP (one time):
  pip install requests openpyxl

CONFIGURE: Fill in the CONFIG section below, then run:
  python dev_build_approver.py
"""

import concurrent.futures
import glob
import os
import time
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import json
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Reports are written to <project-root>/reports/qc/ regardless of where the script is run from.
# Input JSON (from dev_build_creator.py) is also found there automatically.
_REPORTS_DIR = os.path.normpath(
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "reports", "qc")
)

# ─────────────────────────────────────────────
#  CONFIG  ← Fill these in before running
# ─────────────────────────────────────────────
CONFIG = {
    "org":                   "scmdevops",
    "project":               "Leo.TPRM",
    "pat":                   "<YOUR_PAT_HERE>",
    "approval_stage_name":   "DEV: QC Approval",
    "approval_comment":      "Auto-approved by dev_build_approver.py",
    # Leave empty to auto-detect the latest dev_build_report_*.json in output_dir
    "input_json":            "",
    "poll_interval_seconds": 60,
    "poll_timeout_minutes":  60,
    "output_dir":            _REPORTS_DIR,
}
# ─────────────────────────────────────────────

BASE_URL = f"https://dev.azure.com/{CONFIG['org']}/{CONFIG['project']}/_apis"
AUTH = ("", CONFIG["pat"])
HEADERS = {"Content-Type": "application/json"}

_session = requests.Session()
_session.auth = AUTH
_session.headers.update(HEADERS)
_retry = Retry(total=3, backoff_factor=0.5, status_forcelist=[429, 500, 502, 503, 504])
_adapter = HTTPAdapter(pool_connections=20, pool_maxsize=20, max_retries=_retry)
_session.mount("https://", _adapter)

# ── State constants ────────────────────────────────────────────────────────────
STATE_ORDER = {
    "approved": 0,
    "build_failed": 1,
    "timed_out": 2,
    "no_approval_found": 3,
    "api_error": 4,
    "pending": 5,
}
STATE_LABELS = {
    "approved":           "Approved",
    "build_failed":       "Build Failed",
    "timed_out":          "Timed Out",
    "no_approval_found":  "No Gate Found",
    "api_error":          "API Error",
    "pending":            "Still Pending",
}
STATE_EMOJI = {
    "approved":           "✅",
    "build_failed":       "❌",
    "timed_out":          "⏱",
    "no_approval_found":  "⚠️",
    "api_error":          "❌",
    "pending":            "...",
}


# ── HTTP helpers ───────────────────────────────────────────────────────────────

def get(url, params=None):
    params = params or {}
    params["api-version"] = "7.1"
    r = _session.get(url, params=params, timeout=(10, 30))
    r.raise_for_status()
    return r.json()


def patch(url, body, params=None):
    params = params or {}
    params["api-version"] = "7.1"
    r = _session.patch(url, json=body, params=params, timeout=(10, 30))
    r.raise_for_status()
    return r.json()


# ── Input loading ──────────────────────────────────────────────────────────────

def resolve_input_json():
    """Return the JSON file to process — explicit path from CONFIG or latest auto-detected."""
    if CONFIG["input_json"]:
        return CONFIG["input_json"]
    pattern = os.path.join(CONFIG["output_dir"], "dev_build_report_*.json")
    matches = sorted(glob.glob(pattern))
    if not matches:
        raise FileNotFoundError(
            f"No dev_build_report_*.json found in '{CONFIG['output_dir']}'.\n"
            "Run dev_build_creator.py first."
        )
    path = matches[-1]
    print(f"  Using input file: {path}")
    return path


def load_builds(json_path):
    """Load JSON and return only Triggered builds."""
    with open(json_path, encoding="utf-8") as f:
        entries = json.load(f)
    valid = [e for e in entries if e.get("status") == "Triggered" and e.get("build_id")]
    skipped = len(entries) - len(valid)
    if skipped:
        print(f"  Skipped {skipped} non-triggered entries.")
    print(f"  Loaded {len(valid)} triggered build(s).")
    return valid


# ── Azure DevOps polling & approval ───────────────────────────────────────────

def get_build_status(build_id):
    try:
        return get(f"{BASE_URL}/build/builds/{build_id}")
    except Exception as e:
        return {"status": "error", "result": None, "_error": str(e)}


def get_build_timeline(build_id):
    try:
        return get(f"{BASE_URL}/build/builds/{build_id}/timeline").get("records", [])
    except Exception:
        return []


def find_approval_checkpoint(records, stage_name):
    """
    Find the first timeline record that is a waiting/deferred approval checkpoint
    matching stage_name (case-insensitive substring).
    Including 'deferred' allows re-deferral by re-running the script with a new time.
    """
    stage_lower = stage_name.lower()
    for r in records:
        if (
            r.get("type") == "Checkpoint"
            and r.get("state", "").lower() in ("waiting", "pending", "deferred")
            and stage_lower in r.get("name", "").lower()
        ):
            return r
    return None


def approve_checkpoint(approval_id, pipeline_name):
    """Approve the given checkpoint. Returns True on success."""
    try:
        patch(f"{BASE_URL}/pipelines/approvals", [
            {
                "approvalId": approval_id,
                "status":     "approved",
                "comment":    CONFIG["approval_comment"],
            }
        ])
        print(f"    [{pipeline_name}] Approved checkpoint {approval_id}")
        return True
    except requests.HTTPError as e:
        code = e.response.status_code if e.response is not None else "?"
        try:
            detail = e.response.json().get("message", str(e)) if e.response is not None else str(e)
        except Exception:
            detail = str(e)
        print(f"    [{pipeline_name}] FAILED to approve — HTTP {code}: {detail[:120]}")
        return False
    except Exception as e:
        print(f"    [{pipeline_name}] FAILED to approve — {e}")
        return False


# ── Per-build poll logic ───────────────────────────────────────────────────────

def poll_and_approve_build(entry, states, start_times):
    build_id      = entry["build_id"]
    pipeline_name = entry["pipeline_name"]
    state         = states[build_id]

    if state["state"] != "pending":
        return

    # Timeout check
    elapsed = (datetime.now(timezone.utc) - start_times[build_id]).total_seconds()
    if elapsed > CONFIG["poll_timeout_minutes"] * 60:
        state["state"] = "timed_out"
        return

    build_resp = get_build_status(build_id)

    if "_error" in build_resp:
        state["state"] = "api_error"
        state["error"] = build_resp["_error"]
        return

    status = build_resp.get("status", "")
    result = build_resp.get("result")

    if status == "completed":
        if result == "succeeded":
            state["state"] = "no_approval_found"
        else:
            state["state"] = "build_failed"
            state["result"] = result or "unknown"
        return

    # Build is inProgress or notStarted — check timeline for the approval gate
    records    = get_build_timeline(build_id)
    checkpoint = find_approval_checkpoint(records, CONFIG["approval_stage_name"])

    if checkpoint:
        success = approve_checkpoint(checkpoint["id"], pipeline_name)
        if success:
            state["state"]       = "approved"
            state["approved_at"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
        else:
            state["state"] = "api_error"
            state["error"] = f"Approval PATCH failed for checkpoint {checkpoint['id']}"


# ── Poll loop ──────────────────────────────────────────────────────────────────

def print_status_table(entries, states, cycle):
    print(f"\n  {'─'*66}")
    print(f"  Cycle {cycle} — {datetime.now(timezone.utc).strftime('%H:%M:%S UTC')}")
    print(f"  {'─'*66}")
    for entry in entries:
        sid   = entry["build_id"]
        name  = entry["pipeline_name"][:38].ljust(38)
        s     = states[sid]["state"]
        emoji = STATE_EMOJI.get(s, "?")
        label = STATE_LABELS.get(s, s)
        print(f"  {name}  {emoji}  {label}")
    print(f"  {'─'*66}")
    pending  = sum(1 for e in entries if states[e["build_id"]]["state"] == "pending")
    approved = sum(1 for e in entries if states[e["build_id"]]["state"] == "approved")
    failed   = sum(1 for e in entries if states[e["build_id"]]["state"] == "build_failed")
    timedout = sum(1 for e in entries if states[e["build_id"]]["state"] == "timed_out")
    print(f"  Pending: {pending}  |  Approved: {approved}  |  Failed: {failed}  |  Timed Out: {timedout}\n")


def run_poll_loop(entries):
    states = {
        e["build_id"]: {"state": "pending", "approved_at": None, "error": None, "result": None}
        for e in entries
    }
    start_times = {e["build_id"]: datetime.now(timezone.utc) for e in entries}

    print(
        f"  Polling {len(entries)} build(s) — "
        f"every {CONFIG['poll_interval_seconds']}s, "
        f"timeout {CONFIG['poll_timeout_minutes']}min\n"
    )

    cycle = 0
    try:
        while any(s["state"] == "pending" for s in states.values()):
            cycle += 1
            pending_entries = [e for e in entries if states[e["build_id"]]["state"] == "pending"]
            workers = min(len(pending_entries), 8)
            with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as ex:
                futures = {
                    ex.submit(poll_and_approve_build, entry, states, start_times): entry
                    for entry in pending_entries
                }
                concurrent.futures.wait(futures)

            print_status_table(entries, states, cycle)

            if any(s["state"] == "pending" for s in states.values()):
                time.sleep(CONFIG["poll_interval_seconds"])
    except KeyboardInterrupt:
        print("\n\n  Interrupted by user — exiting.")
        os._exit(1)

    return states


# ── Result merging ─────────────────────────────────────────────────────────────

def build_final_results(entries, states):
    results = []
    for entry in entries:
        bid   = entry["build_id"]
        state = states[bid]
        results.append({
            "pipeline_name": entry.get("pipeline_name", ""),
            "status":        state["state"],
            "build_number":  entry.get("build_number", "—"),
            "branch":        entry.get("branch", CONFIG.get("dev_branch", "DEVELOPMENT")),
            "approved_at":   state.get("approved_at") or "—",
            "comment":       CONFIG["approval_comment"] if state["state"] == "approved" else "—",
            "notes":         state.get("error") or state.get("result") or "",
            "build_url":     entry.get("build_url", ""),
            "build_id":      bid,
            "queued_at":     entry.get("queued_at", "—"),
        })
    results.sort(key=lambda r: (STATE_ORDER.get(r["status"], 99), r["pipeline_name"]))
    return results


# ── Excel report ───────────────────────────────────────────────────────────────

def build_excel_report(results):
    wb = Workbook()

    BLUE_DARK  = "1F3864"
    BLUE_MID   = "2E75B6"
    BLUE_LIGHT = "D6E4F0"
    GREEN_BG   = "E2EFDA"
    AMBER_BG   = "FFF2CC"
    RED_BG     = "FFDDD8"
    WHITE      = "FFFFFF"
    GREY_ROW   = "F5F7FA"

    def header_font(size=11, bold=True, color=WHITE):
        return Font(name="Arial", size=size, bold=bold, color=color)

    def cell_font(size=10, bold=False, color="000000"):
        return Font(name="Arial", size=size, bold=bold, color=color)

    def fill(hex_color):
        return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

    def thin_border():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def left():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ── Sheet 1 ────────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "DEV Approval Report"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    ws["A1"] = "QC Release — DEV Build Approver"
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color=WHITE)
    ws["A1"].fill = fill(BLUE_DARK)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    ws["A2"] = (
        f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
        f"  |  Org: {CONFIG['org']}  |  Project: {CONFIG['project']}"
    )
    ws["A2"].font = Font(name="Arial", size=9, color=WHITE)
    ws["A2"].fill = fill(BLUE_MID)
    ws["A2"].alignment = center()
    ws.row_dimensions[2].height = 18

    total    = len(results)
    approved = sum(1 for r in results if r["status"] == "approved")
    failed   = sum(1 for r in results if r["status"] == "build_failed")
    timedout = sum(1 for r in results if r["status"] == "timed_out")

    ws.merge_cells("A3:B3"); ws["A3"] = "Total Builds"
    ws.merge_cells("C3:D3"); ws["C3"] = "Approved"
    ws.merge_cells("E3:F3"); ws["E3"] = "Build Failed"
    ws.merge_cells("G3:H3"); ws["G3"] = "Timed Out"
    ws.merge_cells("A4:B4"); ws["A4"] = total
    ws.merge_cells("C4:D4"); ws["C4"] = approved
    ws.merge_cells("E4:F4"); ws["E4"] = failed
    ws.merge_cells("G4:H4"); ws["G4"] = timedout

    for col_letter, bg in [("A", BLUE_LIGHT), ("C", GREEN_BG), ("E", RED_BG), ("G", AMBER_BG)]:
        for row in [3, 4]:
            cell = ws[f"{col_letter}{row}"]
            cell.fill = fill(bg)
            cell.alignment = center()
            cell.font = cell_font(9, bold=True) if row == 3 else Font(name="Arial", size=20, bold=True)

    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 40
    ws.row_dimensions[5].height = 8

    headers    = ["Repo / Pipeline", "Status", "Build #", "Branch", "Approved At", "Comment", "Notes", "Build Link"]
    col_widths = [35, 20, 14, 20, 22, 30, 30, 14]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=6, column=col_idx, value=h)
        cell.font = header_font()
        cell.fill = fill(BLUE_MID)
        cell.alignment = center()
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[6].height = 22

    STATUS_FILL = {
        "approved":          (fill(GREEN_BG), "276221"),
        "build_failed":      (fill(RED_BG),   "8B0000"),
        "timed_out":         (fill(AMBER_BG), "7D5200"),
        "no_approval_found": (fill(AMBER_BG), "7D5200"),
        "api_error":         (fill(RED_BG),   "8B0000"),
        "pending":           (fill(AMBER_BG), "7D5200"),
    }

    for row_idx, r in enumerate(results, start=7):
        bg = GREY_ROW if row_idx % 2 == 0 else WHITE
        status = r["status"]
        s_fill, s_color = STATUS_FILL.get(status, (fill(RED_BG), "8B0000"))
        link_label = "Open ↗" if status in ("approved", "no_approval_found") else "—"

        row_data = [
            r["pipeline_name"],
            STATE_LABELS.get(status, status),
            r["build_number"],
            r["branch"],
            r["approved_at"],
            r["comment"],
            r["notes"],
            link_label,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border()
            cell.alignment = left() if col_idx in (1, 7) else center()
            cell.font = cell_font()

            if col_idx == 2:
                cell.fill = s_fill
                cell.font = Font(name="Arial", size=10, bold=True, color=s_color)
            else:
                cell.fill = fill(bg)

            if col_idx == 8 and r.get("build_url") and status in ("approved", "no_approval_found"):
                cell.hyperlink = r["build_url"]
                cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")

        ws.row_dimensions[row_idx].height = 18

    ws.freeze_panes = "A7"

    # ── Sheet 2: How To Use ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("How To Use")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 90

    instructions = [
        ("DEV Build Approver — How To Use", BLUE_DARK, 14, True),
        ("", WHITE, 10, False),
        ("PURPOSE", BLUE_MID, 11, True),
        ("  Run this script after dev_build_creator.py to auto-approve the DEV: QC Approval", WHITE, 10, False),
        ("  gate for every triggered build. The approver can then use the Excel report to review.", WHITE, 10, False),
        ("", WHITE, 10, False),
        ("WORKFLOW", BLUE_MID, 11, True),
        ("  1. Run dev_build_creator.py  → triggers builds, saves dev_build_report_*.json", WHITE, 10, False),
        ("  2. Run dev_build_approver.py → polls builds, approves gate, saves approval report", WHITE, 10, False),
        ("  You can run both scripts in separate terminals — the approver starts polling immediately.", WHITE, 10, False),
        ("  If builds are already at the gate, they are approved in the first poll cycle.", WHITE, 10, False),
        ("", WHITE, 10, False),
        ("STEP 1 — Configure the script", BLUE_MID, 11, True),
        ("  • pat                  → Full-access PAT (Build: Read + Pipelines: Read & Manage)", WHITE, 10, False),
        ("  • approval_stage_name  → Name of the gate (default: 'DEV: QC Approval')", WHITE, 10, False),
        ("  • input_json           → Leave empty to auto-detect the latest report file", WHITE, 10, False),
        ("  • poll_interval_seconds→ How often to re-check builds (default: 60)", WHITE, 10, False),
        ("  • poll_timeout_minutes → Give up after this many minutes (default: 60)", WHITE, 10, False),
        ("", WHITE, 10, False),
        ("STEP 2 — Run the script", BLUE_MID, 11, True),
        ("  python dev_build_approver.py", WHITE, 10, False),
        ("", WHITE, 10, False),
        ("UNDERSTANDING THE STATUS COLOURS", BLUE_MID, 11, True),
        ("  Approved        → Gate found and approved successfully", WHITE, 10, False),
        ("  Build Failed    → Build completed with a failure before reaching the gate", WHITE, 10, False),
        ("  Timed Out       → Gate was not reached within poll_timeout_minutes", WHITE, 10, False),
        ("  No Gate Found   → Build completed successfully but no approval gate was seen", WHITE, 10, False),
        ("  API Error       → HTTP or network error during polling or approval", WHITE, 10, False),
    ]

    for i, (text, bg_hex, font_size, bold) in enumerate(instructions, start=1):
        cell = ws2.cell(row=i, column=1, value=text)
        cell.font = Font(name="Arial", size=font_size, bold=bold,
                         color=WHITE if bg_hex != WHITE else "000000")
        cell.fill = fill(bg_hex) if bg_hex != WHITE else PatternFill()
        cell.alignment = left()
        ws2.row_dimensions[i].height = 18

    return wb


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print(f"\n{'='*55}")
    print("  DEV Build Approver — Azure DevOps")
    print(f"{'='*55}")
    print(f"  Org:     {CONFIG['org']}")
    print(f"  Project: {CONFIG['project']}")
    print(f"  Gate:    '{CONFIG['approval_stage_name']}'")
    print(f"  Timeout: {CONFIG['poll_timeout_minutes']} min  |  Interval: {CONFIG['poll_interval_seconds']} s")
    print(f"{'='*55}\n")

    try:
        json_path = resolve_input_json()
    except FileNotFoundError as e:
        print(f"  ERROR: {e}")
        return

    entries = load_builds(json_path)
    if not entries:
        print("  No triggered builds found — nothing to approve.")
        return

    print(f"\n  Starting poll loop for {len(entries)} build(s)...\n")
    states  = run_poll_loop(entries)
    results = build_final_results(entries, states)

    output_dir = CONFIG.get("output_dir", ".")
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
    xlsx_file = os.path.join(output_dir, f"dev_approval_report_{timestamp}.xlsx")
    json_file = os.path.join(output_dir, f"dev_approval_report_{timestamp}.json")

    print("  Building Excel report...")
    wb = build_excel_report(results)
    wb.save(xlsx_file)

    print("  Saving JSON...")
    with open(json_file, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, default=str)

    approved          = sum(1 for r in results if r["status"] == "approved")
    build_failed      = sum(1 for r in results if r["status"] == "build_failed")
    timed_out         = sum(1 for r in results if r["status"] == "timed_out")
    no_approval_found = sum(1 for r in results if r["status"] == "no_approval_found")
    api_errors        = sum(1 for r in results if r["status"] == "api_error")

    print(f"\n{'='*55}")
    print(f"  Reports saved:")
    print(f"       Excel: {xlsx_file}")
    print(f"       JSON:  {json_file}")
    print(f"{'='*55}")
    print(f"  Builds processed  : {len(results)}")
    print(f"  Approved          : {approved}")
    print(f"  Build Failed      : {build_failed}")
    print(f"  Timed Out         : {timed_out}")
    print(f"  No Gate Found     : {no_approval_found}")
    print(f"  API Errors        : {api_errors}")
    print(f"{'='*55}\n")

    if build_failed or timed_out or api_errors:
        print("  Builds that were not approved:")
        for r in results:
            if r["status"] not in ("approved", "no_approval_found"):
                print(f"    - {r['pipeline_name']}: {STATE_LABELS.get(r['status'], r['status'])} — {r['notes']}")
        print()


if __name__ == "__main__":
    main()
